import customtkinter as ctk
import tkinter as tk 
from tkinter import ttk
from PIL import Image
import sys
import re
from tkinter import filedialog
from pathlib import Path
import os, shutil, tempfile
from tkinter import filedialog, messagebox
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import pandas as pd
import matplotlib.pyplot as plt
import threading
from tksheet import Sheet
from tkinter import filedialog
from openpyxl import load_workbook
import mdfreader
from openpyxl.utils import quote_sheetname
from openpyxl.utils import range_boundaries, get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Alignment, Protection
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image, ImageTk, ImageOps,ImageDraw,ImageEnhance
from io import BytesIO

from back import (
    load_signals_from_excel,calculate_time_difference_between_mf4_and_asc,process_all_signals_mdf,calculate_offset_for_synchronization,process_all_signals_mf4,
    fetch_signals_by_frame_name,process_all_signals_of_frame,determine_signal_type,find_error_signals_for_frame,search_signals_in_dsm,
    search_enabled_signals_in_excel,type_of_gateway,identify_multiplexor_signals,extract_basic_receiving_signal,status_table,quality_status_of_signal,plot_communication_sucessful)

class TrieNode:
    def __init__(self):
        self.children = {}
        self.is_end_of_word = False
        self.original_words = set()

class Trie:
    def __init__(self):
        self.root = TrieNode()
        self.all_frames = set()          # All raw frames (for fallback)
        self.root_frames = set()         # Only root-level signal names
        self.full_to_root_map = {}       # Map full names to root signals

    def extract_root_signal(self, signal_name):
        """
        Removes 'DINH' signals and anything after the second underscore.
        For example:
            'DINH_stFId.FId_bInh_Sig435h_Com_uHvbMaxCell_432_ini' → ''
            'SomePrefix_Signal_Extra_Info' → 'SomePrefix_Signal'
        """
        if signal_name.startswith("DINH"):
            return ""

        parts = signal_name.split("_")
        if len(parts) >= 2:
            return "_".join(parts[:2])  # Keep only first two parts
        return signal_name

    def insert(self, word):
        """
        Insert full word but store root signal as primary match.
        Skips 'DINH' and invalid root names.
        """
        self.all_frames.add(word)

        root_word = self.extract_root_signal(word)
        if not root_word:
            return  # Skip inserting 'DINH' or malformed signals

        self.root_frames.add(root_word)
        self.full_to_root_map[word] = root_word

        lower_root = root_word.lower()
        node = self.root

        for char in lower_root:
            if char not in node.children:
                node.children[char] = TrieNode()
            node = node.children[char]

        node.is_end_of_word = True
        node.original_words.add(root_word)

    def search(self, text):
        """
        Search for signals that start with or contain the search text.
        Always returns root signal names only.
        """
        text = text.lower().strip()
        if not text:
            return list(self.root_frames)

        matching_frames = set()

        node = self.root
        for char in text:
            if char not in node.children:
                return self.search_substring(text)
            node = node.children[char]

        matching_frames.update(self._collect_original_words(node))
        matching_frames.update(self.search_substring(text))

        return list(matching_frames)

    def search_substring(self, text):
        """
        Fallback search — substring match in root signal names.
        """
        return [signal for signal in self.root_frames if text in signal.lower()]

    def _collect_original_words(self, node):
        """
        Recursively collect all original root signal names from a node.
        """
        words = list(node.original_words) if node.is_end_of_word else []
        for child_node in node.children.values():
            words.extend(self._collect_original_words(child_node))
        return list(set(words))

    def get_all_frames(self):
        """
        Get all full signals.
        """
        return list(self.all_frames)

    def get_all_root_frames(self):
        """
        Get all deduplicated root signals.
        """
        return list(self.root_frames)

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

class CANwiserApp(ctk.CTk):
    def __init__(self):
        
        super().__init__()
        self.title("CANwiser")
        self.geometry("900x650")
        self.configure(fg_color="#E8DFCA")
        
        self.pages = {}
        self.sync_status = "error"
        self.folder_error_label = None
        self.folder_valid = False
        self.saved_files = {}
        self.temp_dir = tempfile.mkdtemp()

        # Initialize sync labels to None
        self.sync_success_label = ctk.CTkLabel(self, text="")  # Placeholder label
        self.sync_error_label = ctk.CTkLabel(self, text="")    # Placeholder label
        self.function_status = "Idle"
        self.init_main_page()
        
        
        self.clickable_cells = {}  # To track clickable "View Plot" cells
        self.selected_frames = []
        self.window_open = False
        self.current_win = None
        self.session = {}
        self.frames_data = {}
        self.current_frame_index = 0  # Track which frame is being displayed
        self.frame_widgets = []  # Store frame widgets for navigation
        self.frames_to_process = []
        self.saved_files = {}  
        self.frame_id = 1  
        self.dbc_signals = []  
        self.asw_signals = [] 
        self.synchronized_signals_mf4 = []  
        self.enabled_error_signals = [] 
        self.gateway_signals = []  
        self.gateway_types = []  
        self.synchronized_signals_mdf = [] 
        self.basic_receiving_signal = []  
        self.multiplexor_signals = []  
        self.selector_signals = {} 
        self.transmitter_signals = []  
        self.time_shift = 0 
        self.qulaity_analysis_dataframe = None
        self.quality_analysis_page_window = None
        self.row_to_gateway_figs = {}
        self.row_to_rx_figs = {}
        self.frame_name = None
        self.frame_sync_results = {}
        
        self.window_open = False  # Flag to track if the figures window is open
        self.current_win = None
        self.init_analysis_scope_page()
        self.init_sync_page()
        self.init_processing_page()
        self.init_template_page()
        self.show_page("main")
        self.after(100, lambda: self.state("zoomed"))
        self.dynamic_widgets = []
    
    def download_template(self, template_type):
        templates = {
            "can_matrix": "can_matrix_template.xlsx",
            "dtc_matrix": "dtc_matrix_template.xlsx",
            "gateway": "gateway_template.xlsx",
            "fid": "fid_mapping_template.xlsx",
            "report": "report_template.xlsx"}

        template_name = templates.get(template_type)
        if not template_name:
            messagebox.showerror("Error", "Template type not recognized.")
            return

        try:
            # Path to template inside bundled app or dev mode
            base_path = getattr(sys, '_MEIPASS', os.path.abspath("."))
            template_path = os.path.join(base_path, "templates", template_name)

            # Define auto-download location: user's Downloads folder
            downloads_folder = str(Path.home() / "Downloads")
            os.makedirs(downloads_folder, exist_ok=True)
            save_path = os.path.join(downloads_folder, template_name)

            # Copy the file
            shutil.copy(template_path, save_path)
            messagebox.showinfo("Success", f"Template downloaded to:\n{save_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save template:\n{e}")
    
    def init_main_page(self):
        page = ctk.CTkFrame(self, fg_color="#E8DFCA")
        self.pages["main"] = page
        page.pack(fill="both", expand=True)

        page.columnconfigure(0, weight=1)
        page.columnconfigure(1, weight=1)
        page.rowconfigure(0, weight=0)
        page.rowconfigure(1, weight=1)

        # === TOP TITLE ===
        title_label = ctk.CTkLabel(
            page,
            text="WELCOME TO CANWiser",
            font=("Montserrat Black", 60, "bold"),
            text_color="#3E2723",
            anchor="center"
        )
        title_label.place(relx=0.5, y=100, anchor="n")

        # === SUBTITLE ===
        subtitle_label = ctk.CTkLabel(
            page,
            text="Your intelligent companion for CAN signal analysis & reporting.",
            font=("Lato Italic", 20, "italic"),
            text_color="#6D4C41"
        )
        subtitle_label.grid(row=0, column=0, columnspan=2, pady=(180, 0), sticky="n")

        # === LEFT SIDE BUTTONS ===
        left_frame = ctk.CTkFrame(page, fg_color="transparent")
        left_frame.grid(row=1, column=0, sticky="nsew")

        # Make left_frame expand and center its content
        left_frame.grid_rowconfigure(0, weight=1)
        left_frame.grid_rowconfigure(2, weight=1)
        left_frame.grid_columnconfigure(0, weight=1)

        btn_frame = ctk.CTkFrame(left_frame, fg_color="transparent")
        btn_frame.grid(row=0, column=0, sticky="", pady=(50, 0))

        # Pack buttons with pady and center inside btn_frame
        analysis_button = ctk.CTkButton(
            btn_frame,
            text="🚦   Measurement Analysis",
            font=("Inter", 24, "bold"),
            fg_color="#6D4C41",
            hover_color="#7a9669",
            text_color="#FFFFFF",
            corner_radius=12,
            height=80,
            width=500,
            command=lambda: self.set_function_status("Measurement Analysis")
        )
        analysis_button.pack(pady=(10,10), anchor="center") 

        report_button = ctk.CTkButton(
            btn_frame,
            text="📄  Generate Report",
            font=("Inter", 24, "bold"),
            fg_color="#6D4C41",
            hover_color="#7a9669",
            text_color="#FFFFFF",
            corner_radius=12,
            height=80,
            width=500,
            command=lambda: self.set_function_status("Report Generation")
        )
        report_button.pack(pady=(25, 10), anchor="center")

        right_frame = ctk.CTkFrame(page, fg_color="transparent")
        right_frame.grid(row=1, column=1, sticky="nsew")
        right_frame.grid_rowconfigure(0, weight=1)
        right_frame.grid_columnconfigure(0, weight=1)

        # === Card Frame ===
        card_frame = ctk.CTkFrame(right_frame, fg_color="transparent")
        card_frame.grid(row=0, column=0, pady=(40, 20), padx=10, sticky="n")

        canvas_width, canvas_height = 860, 620
        self.canvas = tk.Canvas(
            card_frame, width=canvas_width, height=canvas_height,
            bg="#E8DFCA", bd=0, highlightthickness=0, relief='ridge'
        )
        self.canvas.pack()

        self.card_bg_img = self.draw_rounded_card(canvas_width, canvas_height)
        self.canvas.create_image(0, 0, anchor="nw", image=self.card_bg_img)

        # === Load and Prepare Images ===
        image_files = ["frame1.png", "frame2.png", "frame3.png"]
        self.images_color = [Image.open(f).convert("RGBA") for f in image_files]

        # Add gray border to all
        self.images_color = [
            ImageOps.expand(img, border=10, fill="#BDBDBD") for img in self.images_color
        ]

        # Center images: Full-color
        self.images_center = [
            img.resize((600, 400), Image.Resampling.LANCZOS) for img in self.images_color
        ]

        # Side images: Grayscale + faded
        self.images_side = []
        for img in self.images_color:
            gray = ImageOps.grayscale(img.resize((500, 300), Image.Resampling.LANCZOS)).convert("RGBA")
            faded = ImageEnhance.Brightness(gray).enhance(0.3)
            self.images_side.append(faded)

        self.current_index = 0
        self.image_refs = {}

        # === Dot Indicators ===
        dot_frame = ctk.CTkFrame(right_frame, fg_color="transparent")
        dot_frame.place(x=400, y=450)
        self.dots = []
        for i in range(len(self.images_color)):
            dot = ctk.CTkLabel(dot_frame, text="●", text_color="#BDBDBD", font=("Arial", 16))
            dot.pack(side="left", padx=4)
            self.dots.append(dot)
        self.dots[0].configure(text_color="#212121")

        self.draw_carousel()
        self.after(3000, self.animate_transition)

    def draw_rounded_card(self, width, height, radius=30, color="#E8DFCA"):
        img = Image.new("RGBA", (width, height), (0, 0, 0, 0))
        draw = ImageDraw.Draw(img)
        draw.rounded_rectangle([(0, 0), (width, height)], radius=radius, fill=color)
        return ImageTk.PhotoImage(img)

    def draw_carousel(self):
        self.canvas.delete("carousel")
        idx = self.current_index
        total = len(self.images_color)

        left = (idx - 1) % total
        center = idx
        right = (idx + 1) % total

        img_left = ImageTk.PhotoImage(self.images_side[left])
        self.canvas.create_image(50, 100, anchor="nw", image=img_left, tags="carousel")

        img_right = ImageTk.PhotoImage(self.images_side[right])
        self.canvas.create_image(400, 100, anchor="nw", image=img_right, tags="carousel")

        self.canvas.create_rectangle(
            140, 50, 760, 470,
            fill="#D6CEC2", outline="", tags="carousel"
        )

        img_center = ImageTk.PhotoImage(self.images_center[center])
        self.canvas.create_image(150, 60, anchor="nw", image=img_center, tags="carousel")

        self.image_refs = {
            "left": img_left,
            "center": img_center,
            "right": img_right
        }

        self.update_dots()

    def animate_transition(self):
        steps = 15
        delay = 20  # milliseconds

        idx = self.current_index
        total = len(self.images_color)
        left = (idx - 1) % total
        center = idx
        right = (idx + 1) % total
        next_right = (idx + 2) % total

        # Prepare images
        center_img = self.images_center[center]
        right_img = self.images_side[right]
        next_img = self.images_side[next_right]

        for step in range(steps + 1):
            self.canvas.delete("carousel")
            t = step / steps

            # Center shrinking and moving left
            c_w = int(600 - 100 * t)
            c_h = int(400 - 100 * t)
            c_x = int(150 - 100 * t)
            c_y = int(60 + 40 * t)
            img_c = ImageTk.PhotoImage(center_img.resize((c_w, c_h)))

            # Right enlarging and moving center
            r_w = int(500 + 100 * t)
            r_h = int(300 + 100 * t)
            r_x = int(400 - 250 * t)
            r_y = int(100 - 40 * t)
            img_r = ImageTk.PhotoImage(right_img.resize((r_w, r_h)))

            # Next image moving in
            n_x = int(800 - 400 * t)
            img_n = ImageTk.PhotoImage(next_img.resize((500, 300)))

            self.canvas.create_rectangle(
                140, 50, 760, 470,
                fill="#E8DFCA", outline="", tags="carousel"
            )
            self.canvas.create_image(c_x, c_y, anchor="nw", image=img_c, tags="carousel")
            self.canvas.create_image(r_x, r_y, anchor="nw", image=img_r, tags="carousel")
            self.canvas.create_image(n_x, 100, anchor="nw", image=img_n, tags="carousel")

            self.image_refs = {"c": img_c, "r": img_r, "n": img_n}
            self.update()
            self.after(delay)

        self.current_index = (self.current_index + 1) % len(self.images_color)
        self.draw_carousel()
        self.after(1000, self.animate_transition)

    def update_dots(self):
        for i, dot in enumerate(self.dots):
            dot.configure(text_color="#212121" if i == self.current_index else "#BDBDBD")

    def set_function_status(self, status):
        self.function_status = status
        print("Function Status Set:", self.function_status)

        # Always recreate the measurement page based on function status
        self.pages.pop("measurement", None)  # Remove existing page if present
        self.init_measurement_page()         # Reinitialize fresh with correct widgets

        self.show_page("measurement") 
    
    
    # -----------------------------------------------------------------------------------------------------
    #----------------------------------------------------------------------------------------------------------
    
    
    
    def init_measurement_page(self):
        page = ctk.CTkFrame(self, fg_color="#E8DFCA")
        self.pages["measurement"] = page
        
        if self.function_status == "Measurement Analysis":
            ctk.CTkLabel(page, text="Measurement Analysis", font=ctk.CTkFont("Arial", 40, "bold"), text_color="#3E2723").pack(pady=(50, 20),anchor="center")
        elif self.function_status == "Report Generation":
            ctk.CTkLabel(page, text="Report Generation", font=ctk.CTkFont("Arial", 40, "bold"), text_color="#3E2723").pack(pady=(50, 20),anchor="center")
        
        ctk.CTkLabel(page, text="Select the Input Sheets", font=ctk.CTkFont("Arial", 20), text_color="#6D4C41").pack(pady=(10, 20),anchor="center")

        frame = ctk.CTkFrame(page, fg_color="transparent")
        frame.pack(pady=5, anchor="w", padx=30)
        
        self.can_entry = ctk.CTkEntry(frame, width=500, height=40, placeholder_text="Select CAN Matrix")
        self.can_entry.grid(row=0, column=1, padx=10, pady=15)
        ctk.CTkButton(frame, text="📄", width=40,fg_color="#D7CCC8",hover_color="#BCAAA4",text_color="#4E342E",  command=self.select_can).grid(row=0, column=2, padx=5)
        ctk.CTkLabel(frame, text="CAN MATRIX", font=("Arial", 18 , "bold"), text_color="#6D4C41").grid(row=0, column=0, sticky="w", padx=(200, 5))
        ctk.CTkButton(frame, text="📥 Download CAN Matrix Template", width=40,fg_color="#8D6E63",hover_color="#795548",text_color="#FFFFFF",  
                      command=lambda: self.download_template("can_matrix")).grid(row=0, column=3, padx=10)


        self.dtc_entry = ctk.CTkEntry(frame, width=500, height=40, placeholder_text="Select DTC Matrix")
        self.dtc_entry.grid(row=1, column=1, padx=50, pady=5)
        ctk.CTkButton(frame, text="📄", width=40,fg_color="#D7CCC8",hover_color="#BCAAA4",text_color="#4E342E", command=self.select_dtc).grid(row=1, column=2, padx=5)
        ctk.CTkLabel(frame, text="DTC MATRIX", font=("Arial", 18,"bold"), text_color="#6D4C41").grid(row=1, column=0, sticky="w", padx=(200, 5))
        ctk.CTkButton(frame, text="📥 Download DTC Matrix Template", width=40,fg_color="#8D6E63",hover_color="#795548",text_color="#FFFFFF",
                       command=lambda: self.download_template("dtc_matrix")).grid(row=1, column=3, padx=10)

        self.gateway_entry = ctk.CTkEntry(frame, width=500, height=40, placeholder_text="Select Gateway Sheet") 
        self.gateway_entry.grid(row=2, column=1, padx=50, pady=15)
        ctk.CTkButton(frame, text="📄", width=40,fg_color="#D7CCC8",hover_color="#BCAAA4",text_color="#4E342E", command=self.select_gateway).grid(row=2, column=2, padx=5)
        ctk.CTkLabel(frame, text="GATEWAY SHEET", font=("Arial", 18,"bold"), text_color="#6D4C41").grid(row=2, column=0, sticky="w", padx=(200, 5))
        ctk.CTkButton(frame, text="📥 Download Gateway Sheet Template", width=40,fg_color="#8D6E63",hover_color="#795548",text_color="#FFFFFF",
                       command=lambda: self.download_template("gateway")).grid(row=2, column=3, padx=10)

        self.fid_entry = ctk.CTkEntry(frame, width=500, height=40, placeholder_text="Select FID Mapping")
        self.fid_entry.grid(row=3, column=1, padx=50, pady=15)
        ctk.CTkButton(frame, text="📄", width=40,fg_color="#D7CCC8",hover_color="#BCAAA4",text_color="#4E342E", command=self.select_fid).grid(row=3, column=2, padx=5)
        ctk.CTkLabel(frame, text="FID MAPPING", font=("Arial", 18,"bold"), text_color="#6D4C41").grid(row=3, column=0, sticky="w", padx=(200, 5))
        ctk.CTkButton(frame, text="📥 Download Fid Mapping Template", width=40,fg_color="#8D6E63",hover_color="#795548",text_color="#FFFFFF",
                       command=lambda: self.download_template("fid")).grid(row=3, column=3, padx=10)
        

        self.folder_entry = ctk.CTkEntry(frame, width=500, height=40, placeholder_text="Select Measurement Folder")
        self.folder_entry.grid(row=4, column=1, padx=50, pady=15)
        ctk.CTkButton(frame, text="📁", width=40,fg_color="#D7CCC8",hover_color="#BCAAA4",text_color="#4E342E", command=self.select_folder).grid(row=4, column=2, padx=5)
        ctk.CTkLabel(frame, text="MEASUREMENTS FOLDER", font=("Arial", 18,"bold"), text_color="#6D4C41").grid(row=4, column=0, sticky="w", padx=(200, 5))
        
        
        if self.function_status.lower() == "report generation":
            self.report_entry = ctk.CTkEntry(frame, width=500,height=40, placeholder_text="Select Report Tempelate")
            self.report_entry.grid(row=5, column=1, padx=50, pady=15)
            ctk.CTkButton(frame, text="📄", width=40,fg_color="#D7CCC8",hover_color="#BCAAA4",text_color="#4E342E",
                           command=self.select_report_tempelate).grid(row=5, column=2, padx=5)
            ctk.CTkLabel(frame, text="REPORT TEMPLATE", font=("Arial", 18,"bold"), text_color="#6D4C41").grid(row=5, column=0, sticky="w", padx=(200, 5))
            ctk.CTkButton(frame, text="📥 Download Report Template", width=40,fg_color="#8D6E63",hover_color="#795548",text_color="#FFFFFF",
                           command=lambda: self.download_template("report")).grid(row=5, column=3, padx=10)

        self.drop_menu = ctk.CTkOptionMenu(page,values=["CANalyser Plot/Recording", "HIL recording", "CAN monitoring/recording"],fg_color="#8D6E63",button_color="#D9B382",
                                           button_hover_color="#D9B382",text_color="#FFFFFF")
        self.drop_menu.set("Select Type of Recording")
        self.drop_menu.pack(pady=(30, 10))

        # Next Button in warm beige
        ctk.CTkButton(page,text="Next",width=120,fg_color="#C9A175",hover_color="#D9B382",text_color="#4E342E",command=self.check_selection_and_proceed).pack(pady=10)

        # Back Button styled the same
        ctk.CTkButton(page, text="Back", width=120, fg_color="#C9A175", hover_color="#D9B382", text_color="#4E342E", 
                      command=lambda: self.show_page("main")).pack(pady=(10, 30))

        self.add_footer(page)
    
    def check_selection_and_proceed(self):
        selected = self.drop_menu.get()
        if selected == "CANalyser Plot/Recording":
            self.selected_type_of_recording = selected
            self.validate_and_go_to_sync()
        if selected == "CAN monitoring/recording":
            self.selected_type_of_recording = selected
            self.validate_and_go_to_sync()
        else:
            messagebox.showwarning("Selection Required","Only 'CANalyser Plot/Recording' is currently supported. Please select it to proceed.")
    
    # -----------------------------------------------------------------------------------------------------
    #----------------------------------------------------------------------------------------------------------

    def init_template_page(self):
        page = ctk.CTkFrame(self, fg_color="#E8DFCA")
        self.pages["type_of_project"] = page

        # Page Heading
        heading = ctk.CTkLabel(
            page,
            text="Select the Labels Format",
            font=ctk.CTkFont("Arial", 40, "bold"),
            text_color="#3E2723"
        )
        heading.pack(pady=(50, 20), anchor="center")

        # Frame to hold all label types and dropdowns
        dropdown_frame = ctk.CTkFrame(page, fg_color="transparent")
        dropdown_frame.pack(pady=30, anchor="w", padx=100)

        self.label_types = ["onfail", "init_type", "invld"]

        self.known_templates = {
            "onfail": ["{asw_signal}_{frame_id}hRcf_C", "{asw_signal}Dfl_C", "{asw_signal}Def_C"],
            "init_type": ["{asw_signal}_IniTyp_C", "{asw_signal}Ini_Type"],
            "invld": ["{asw_signal}_Invld_C", "{asw_signal}Invalid"]
        }

        self.selected_templates_by_type = {}
        self.template_entries = {}

        for row_idx, label_type in enumerate(self.label_types):
            label_text = self.format_label_type_display(label_type)
            
            ctk.CTkLabel(dropdown_frame, text="").grid(row=row_idx, column=0, padx=20)
            # Label
            ctk.CTkLabel(
                dropdown_frame,
                text=label_text,
                font=ctk.CTkFont("Arial", 18, "bold"),
                text_color="#6D4C41"
            ).grid(row=row_idx, column=1, sticky="w", padx=300)

            # Dropdown (ComboBox)
            entry = ctk.CTkComboBox(
                dropdown_frame,
                values=self.known_templates.get(label_type, []),
                fg_color="#8D6E63",
                button_color="#D9B382",
                button_hover_color="#D9B382",
                text_color="#FFFFFF",
                width=350,height=40
            )
            entry.set("")
            entry.grid(row=row_idx, column=1, padx=600, pady=20)

            self.template_entries[label_type] = entry

        # Navigation Buttons at bottom (still using grid)
        button_frame = ctk.CTkFrame(page, fg_color="transparent")
        button_frame.pack(pady=30)

        ctk.CTkButton(
            button_frame,
            text="Next",
            width=120,
            fg_color="#C9A175",
            hover_color="#D9B382",
            text_color="#4E342E",
            command=self.validate_and_store_all_templates
        ).grid(row=0, column=1, padx=20,pady=30)

        ctk.CTkButton(
            button_frame,
            text="Back",
            width=120,
            fg_color="#C9A175",
            hover_color="#D9B382",
            text_color="#4E342E",
            command=lambda: self.show_page("measurement")
        ).grid(row=0, column=0, padx=20,pady=30)

        self.add_footer(page)



    def format_label_type_display(self, label_type):
        mapping = {
            "onfail": "OnFail/FailSafe Calibration",
            "init_type": "Init_Type Calibration",
            "invld": "Invalid Calibration"
        }
        return mapping.get(label_type, label_type.capitalize())

    def validate_and_store_all_templates(self):
        for label_type, entry in self.template_entries.items():
            selected_template = entry.get().strip()

            if not selected_template:
                print(f"No format entered for {label_type}")
                continue

            self.selected_templates_by_type[label_type] = selected_template

            # Optionally update known list
            if selected_template not in self.known_templates[label_type]:
                self.known_templates[label_type].append(selected_template)

        self.show_page("analysis_scope")

        
    # -----------------------------------------------------------------------------------------------------
    #----------------------------------------------------------------------------------------------------------
    
    
    def init_analysis_scope_page(self):
        page = ctk.CTkFrame(self, fg_color="#E8DFCA")
        self.pages["analysis_scope"] = page

        # Initialize dynamic widgets list
        self.dynamic_widgets = []

        box = ctk.CTkFrame(page,fg_color="#8D6E63",corner_radius=9,width=1500,)
        box.pack(pady=80, padx=20, anchor="center")  # or "w" or "e" for left/right

        can_matrix_path = self.saved_files.get("can_matrix", "")

        # Question Label
        question_label = ctk.CTkLabel(
            box,
            text="Analysis required for all the Signals present in the\nmeasurement/CAN Matrix?",
            font=ctk.CTkFont("Arial", 24, "bold"),
            text_color="#FFFFFF",
            justify="center"
        )
        question_label.pack(pady=(20, 15), padx=20)  # add horizontal padding inside

        # Button frame
        btn_frame = ctk.CTkFrame(box, fg_color="transparent", width=1400)
        btn_frame.pack(pady=10)

        self.analysis_status = "Yes"

        # YES button
        yes_btn = ctk.CTkButton(btn_frame, text="Yes", width=100, fg_color="#fff", text_color="#000",
            hover_color="#ccc", command=lambda: self.full_signal_analysis(page))
        yes_btn.pack(side="left", padx=20)

        # NO button
        no_btn = ctk.CTkButton(btn_frame, text="No", width=100, fg_color="#fff", text_color="#000",
            hover_color="#ccc", command=lambda: self.reveal_frame_input(page))
        no_btn.pack(side="right", padx=20)

        # Back and Next Buttons
        back_btn = ctk.CTkButton(page, text="Back", width=120, fg_color="#C9A175", hover_color="#D9B382", text_color="#4E342E", command=lambda: self.show_page("measurement"))
        back_btn.place(x=40, rely=1.0, y=-50, anchor="sw")

        next_btn = ctk.CTkButton(page, text="Next", width=120, fg_color="#C9A175", hover_color="#D9B382", text_color="#4E342E", command=lambda: self.show_page("sync"))
        next_btn.place(relx=1.0, rely=1.0, x=-40, y=-50, anchor="se")

        self.add_footer(page)

    
    def full_signal_analysis(self, page):
        self.analysis_status = "Yes"

        # Destroy only previously created dynamic widgets (not question_label/buttons)
        for widget in getattr(self, "dynamic_widgets", []):
            widget.destroy()
        self.dynamic_widgets.clear()
        
        can_matrix_path = self.saved_files.get("can_matrix", "")
        if os.path.exists(can_matrix_path):
            self.tx_df, self.rx_df = load_signals_from_excel(can_matrix_path)
            if self.tx_df is not None and self.rx_df is not None:
                self.frame_names = self.extract_frame_names()
        
        self.selected_frames = self.frame_names.copy()
        
        # === New Section ===
        output_section = ctk.CTkFrame(page, fg_color="#E8DFCA", corner_radius=10)
        output_section.pack(pady=10, padx=20, fill="x")
        self.dynamic_widgets.append(output_section)

        ctk.CTkLabel(output_section,text="CAN Frame Overview",
            font=ctk.CTkFont("Arial", 15, weight="bold"),text_color="#3E3E3E").pack(fill="x", pady=(10, 10))

        container = ctk.CTkFrame(output_section, fg_color="#F5F1E6", corner_radius=12)
        container.pack(padx=15, pady=(0, 15))
        self.dynamic_widgets.append(container)

        canvas = tk.Canvas(container, width=1500, height=400, highlightthickness=0)
        canvas.pack(side="left")
        vscrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        vscrollbar.pack(side="right", fill="y")
        canvas.configure(yscrollcommand=vscrollbar.set)

        frame_inside_canvas = ctk.CTkFrame(canvas, fg_color="#F5F1E6")
        canvas.create_window((0, 0), window=frame_inside_canvas, anchor="nw")

        # Store dynamically created labels
        for i, frame in enumerate(self.frame_names):
            row = i // 6
            col = i % 6
            label = ctk.CTkLabel(frame_inside_canvas, text=frame, font=ctk.CTkFont("Arial", 11),
                                text_color="#4B4B4B", fg_color="#DACBBE", corner_radius=14, padx=8, pady=3)
            label.grid(row=row, column=col, padx=4, pady=4, sticky="w")
            self.dynamic_widgets.append(label)

        def on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        frame_inside_canvas.bind("<Configure>", on_frame_configure)

        if hasattr(self, "suggestion_frame"):
            self.suggestion_frame.pack_forget()
    
    def reveal_frame_input(self, page):
        # Destroy existing dynamic widgets (if any)
        for widget in getattr(self, "dynamic_widgets", []):
            widget.destroy()
        self.dynamic_widgets.clear()

        self.analysis_status = "No"

        # Recreate frame_input_section
        self.frame_input_section = ctk.CTkFrame(page, fg_color="transparent")
        self.frame_input_section.pack(pady=20)
        self.dynamic_widgets.append(self.frame_input_section)

        label_title = ctk.CTkLabel(
            self.frame_input_section, text="🔍 Input Frame(s) for Analysis.",
            font=ctk.CTkFont("Arial", 24, "bold"), text_color="#3E2723"
        )
        label_title.pack(pady=(10, 10))
        self.dynamic_widgets.append(label_title)

        content_frame = ctk.CTkFrame(self.frame_input_section, fg_color="transparent")
        content_frame.pack(pady=10)
        self.dynamic_widgets.append(content_frame)

        can_matrix_path = self.saved_files.get("can_matrix", "")
        if os.path.exists(can_matrix_path):
            try:
                self.tx_df, self.rx_df = load_signals_from_excel(can_matrix_path)
                if self.tx_df is not None and self.rx_df is not None:
                    self.trie = Trie()
                    self.frame_names = self.extract_frame_names()
                    for frame_name in self.frame_names:
                        self.trie.insert(frame_name)

                    label_input = ctk.CTkLabel(content_frame,text="Enter the frame for Analysis:\nPlease Press Enter to see the selected frames.",
                        font=ctk.CTkFont("Arial", 18),text_color="#6D4C41", anchor="e", width=220)
                    label_input.grid(row=0, column=0, padx=(20, 10), pady=10, sticky="e")
                    self.dynamic_widgets.append(label_input)

                    self.frame_input = ctk.CTkEntry(content_frame, width=500, height=60,font=ctk.CTkFont("Arial", 14), justify="left")
                    self.frame_input.grid(row=0, column=1, padx=(0, 20), pady=10)
                    self.dynamic_widgets.append(self.frame_input)

                    self.frame_input.bind("<KeyRelease>", self.search_and_suggest_frames)
                    self.frame_input.bind("<Return>", self.display_selected_frames)
                    self.frame_input.bind("<FocusIn>", self.search_and_suggest_frames)

                    self.suggestion_frame = ctk.CTkFrame(page, width=500, height=200, fg_color="transparent")
                    self.suggestion_frame.pack(pady=(5, 10), padx=10, anchor="center")
                    self.suggestion_frame.pack_forget()
                    self.dynamic_widgets.append(self.suggestion_frame)

                    label_output = ctk.CTkLabel(
                        content_frame, text="Your selected Frame(s) are:",
                        font=ctk.CTkFont("Arial", 18), text_color="#6D4C41",
                        anchor="e", width=220)
                    label_output.grid(row=1, column=0, padx=(10, 10), pady=15, sticky="e")
                    self.dynamic_widgets.append(label_output)

                    self.selected_frames_display = ctk.CTkLabel(content_frame, text="", font=ctk.CTkFont("Arial", 18), text_color="#6D4C41")
                    self.selected_frames_display.grid(row=1, column=1, padx=(0, 20), pady=10, sticky="w")
                    self.dynamic_widgets.append(self.selected_frames_display)
            
            except Exception as e:
                self.sync_error_label = ctk.CTkLabel(page,text=f"❌ Error loading CAN Matrix: {str(e)}",font=("Arial", 14, "bold"),fg_color="#FF6B6B",
                    text_color="black",width=600)
                self.sync_error_label.pack(pady=(10, 10))
                self.dynamic_widgets.append(self.sync_error_label)
    
    def extract_frame_names(self):
        tx_signals_frame = self.tx_df['Frame Name'].dropna().unique().tolist() if 'Frame Name' in self.tx_df.columns else []
        rx_signals_frame = self.rx_df['Frame Name'].dropna().unique().tolist() if 'Frame Name' in self.rx_df.columns else []

        return list(set(tx_signals_frame + rx_signals_frame))
    
    def display_selected_frames(self, event=None):
        if self.suggestion_frame:
            # Just hide suggestion box and clear suggestions inside
            self.suggestion_frame.pack_forget()
            self.destroy_suggestions()

        user_input = self.frame_input.get()
        if user_input:
            # Split and clean the input
            new_frames = [frame.strip() for frame in user_input.split(",") if frame.strip()]

            # Clear existing selected frames
            self.selected_frames = []

            # Add new frames
            for frame in new_frames:
                self.selected_frames.append(frame)

            # Update the display
            display_text = ", ".join(self.selected_frames)
            self.selected_frames_display.configure(text=display_text)

        # Remove focus from entry to prevent immediate re-showing of suggestions
        self.frame_input.master.focus()
    
    def search_and_suggest_frames(self, event=None):
        typed_text = self.frame_input.get().strip()

        if not typed_text:
            if self.suggestion_frame:
                self.suggestion_frame.pack_forget()
            return

        frames = [frame.strip() for frame in typed_text.split(",")]
        current_frame = frames[-1] if frames else ""

        if not current_frame:
            if self.suggestion_frame:
                self.suggestion_frame.pack_forget()
            return

        matching_frames = [frame for frame in self.trie.get_all_frames()
                        if current_frame.lower() in frame.lower()]

        # Clear previous suggestions first
        self.destroy_suggestions()

        if matching_frames:
            for frame in matching_frames:
                frame_option = ctk.CTkButton(
                    self.suggestion_frame,
                    text=frame,
                    fg_color="#E0E0E0",
                    hover_color="#D0D0D0",
                    text_color="#000000",
                    font=("Arial", 12),
                    command=lambda s=frame: self.add_frame_to_input(s)
                )
                frame_option.pack(padx=10, anchor="center")
        else:
            no_match_frame = ctk.CTkLabel(
                self.suggestion_frame,
                text="No matches found",
                font=("Arial", 12),
                text_color="gray"
            )
            no_match_frame.pack(padx=10, anchor="center")

        # Show suggestion frame if hidden
        if self.suggestion_frame and not self.suggestion_frame.winfo_ismapped():
            self.suggestion_frame.pack(pady=(5, 10), padx=10, anchor="center")



    def add_frame_to_input(self, frame_name):
        """Add the selected frame to the input field, maintaining the comma-separated list."""
        typed_text = self.frame_input.get().strip()
        frames = [frame.strip() for frame in typed_text.split(",")]

        # Replace the last frame with the selected one
        frames[-1] = frame_name

        # Update the frame input
        self.frame_input.delete(0, "end")
        self.frame_input.insert(0, ", ".join(frames))
        self.suggestion_frame.pack_forget()

    def destroy_suggestions(self):
        """Clear existing suggestions from the suggestion frame."""
        if self.suggestion_frame is not None:
            for widget in self.suggestion_frame.winfo_children():
                widget.destroy()
    
    
    # -------------------------------------SYNCHING PAGE FOR CANALYZER PLOT----------------------------------------------------------------
    #----------------------------------------------------------------------------------------------------------
  
    def init_sync_page(self):
        self.frame_sync_results = {}
        """Initialize the Synchronization page."""
        page = ctk.CTkFrame(self, fg_color="#E8DFCA")
        self.pages["sync"] = page

        ctk.CTkLabel(page, text="Synchronization of the files for Analysis", 
                    font=ctk.CTkFont("Arial", 40, "bold"), text_color="#3E2723").pack(pady=(80, 20))

        scrollable_frame = ctk.CTkScrollableFrame(page, width=900, height=500, fg_color="#E8DFCA")
        scrollable_frame.pack(pady=10, padx=20, fill="x")

        can_matrix_path = self.saved_files.get("can_matrix", "")

        processed_subfolders = {}  # Track processed subfolders for reuse
        self.tries = {}
        if os.path.exists(can_matrix_path):
            try:
                self.tx_df, self.rx_df = load_signals_from_excel(can_matrix_path)
                if self.tx_df is not None and self.rx_df is not None:
                    self.sync_success_label = ctk.CTkLabel(
                        scrollable_frame, text="✅ CAN Matrix Loaded Successfully!", font=("Arial", 14, "bold"),
                        fg_color="#8D9440", text_color="black", width=600)
                    self.sync_success_label.pack(pady=(10, 10))

                if self.analysis_status == "No" or self.analysis_status == "Yes":
                    self.signal_refs = {}
                    self.signal_files = {}
                    folder_path_str = self.folder_entry.get().strip()
                    folder = Path(folder_path_str)

                    # Step 1: Group frames by their subfolder
                    subfolder_to_frames = {}

                    for frame in self.selected_frames:
                        match = re.search(r'(\d{2,})(h)?$', frame)
                        if not match:
                            continue

                        frame_id = match.group(1)
                        print(frame_id)
                        found_subfolder = None
                        fallback_used = False

                        if folder.exists() and folder.is_dir():
                            for subfolder in folder.iterdir():
                                if subfolder.is_dir() and subfolder.name == frame_id:
                                    found_subfolder = subfolder
                                    break
                            if not found_subfolder:
                                general_folder = folder / "general"
                                if general_folder.exists() and general_folder.is_dir():
                                    found_subfolder = general_folder
                                    fallback_used = True

                        if not found_subfolder:
                            no_folder_label = ctk.CTkLabel(scrollable_frame, 
                                text=f"❌ No subfolder found for frame ID: {frame_id} (frame: {frame})",
                                font=("Arial", 12, "italic"), text_color="red")
                            no_folder_label.pack(pady=(5, 5))
                            continue

                        subfolder_key = str(found_subfolder.resolve())
                        subfolder_to_frames.setdefault(subfolder_key, {"path": found_subfolder, "frames": []})
                        subfolder_to_frames[subfolder_key]["frames"].append(frame)
                    
                    # Step 2: Process each unique subfolder only once
                    for subfolder_key, data in subfolder_to_frames.items():
                        found_subfolder = data["path"]
                        frames = data["frames"]

                        frame_container = ctk.CTkFrame(scrollable_frame, fg_color="#EFEBE9", corner_radius=10)
                        frame_container.pack(pady=10, padx=10, fill="x")

                        lines_of_frames = [frames[i:i+18] for i in range(0, len(frames), 8)]

                        # Title label
                        ctk.CTkLabel(frame_container,
                            text="🟫 Frames:",
                            font=("Arial", 16, "bold"),
                            text_color="#4E342E",
                            anchor="w"
                        ).pack(pady=(10, 2), padx=15, anchor="w")

                        # Each line is one label
                        for line in lines_of_frames:
                            frame_line_text = ", ".join(line)
                            ctk.CTkLabel(frame_container,
                                text=frame_line_text,
                                font=("Arial", 12),
                                text_color="#4E342E",
                                anchor="w"
                            ).pack(pady=(0, 2), padx=30, anchor="w")
                        
                        ctk.CTkLabel(
                            frame_container,
                            text=f"📁 Subfolder Used: {found_subfolder.name}",
                            font=("Arial", 14),
                            text_color="#5D4037",
                            anchor="w"
                        ).pack(pady=(0, 8), padx=15, anchor="w")

                        mdf = next(found_subfolder.glob("*.mdf"), None)
                        mf4 = next(found_subfolder.glob("*.mf4"), None)
                        asc = next(found_subfolder.glob("*.asc"), None)

                        for frame in frames:
                            self.signal_files[frame] = {
                                "subfolder": str(found_subfolder),
                                "mdf_file": str(mdf) if mdf else None,
                                "mf4_file": str(mf4) if mf4 else None,
                                "asc_file": str(asc) if asc else None,
                            }

                        if subfolder_key in processed_subfolders:
                            stored = processed_subfolders[subfolder_key]
                            for frame in frames:
                                self.frame_sync_results[frame] = {
                                    "subfolder": subfolder_key,
                                    "mf4_file": stored["mf4_file"],
                                    "mdf_file": stored["mdf_file"],
                                    "asc_file": stored["asc_file"],
                                    "offset": stored["offset"],
                                    "time_shift" : stored["time_shift"],
                                    "plot_fig1": stored["plot_fig1"],
                                    "plot_fig2": stored["plot_fig2"],
                                    "synchronized_signals_mdf": stored.get("synchronized_signals_mdf"),
                                    "synchronized_signals_mf4": stored.get("synchronized_signals_mf4"),
                                }

                            reused_label = ctk.CTkLabel(frame_container,
                                text=f"✅ Reused processed data. Offset: {stored['offset']:.2f} sec",
                                font=("Arial", 13, "bold"), text_color="green")
                            reused_label.pack(padx=15, pady=(5, 5), anchor="w")

                            for frame in frames:
                                ctk.CTkButton(frame_container, text=f"View Plot ({frame})", width=150, fg_color="#007acc",
                                    hover_color="#005f99", text_color="white",
                                    command=lambda f=frame: self.show_plot(f)).pack(padx=15, pady=(2, 5), anchor="w")
                            continue
                        
                        signal_names = self.extract_signal_names_for_subfolder(found_subfolder)
                        trie = Trie()
                        for name in signal_names:
                            trie.insert(name)
                        self.tries[subfolder_key] = trie


                        # Only one entry per subfolder
                        entry_label = ctk.CTkLabel(
                            frame_container,
                            text=f"Enter reference signal for subfolder: {found_subfolder.name}",
                            font=("Arial", 13, "italic"),
                            text_color="#3E2723",
                            anchor="w"
                        )
                        entry_label.pack(padx=15, anchor="w")

                        signal_entry = ctk.CTkEntry(frame_container, width=500, fg_color="#EFEBE9", text_color="#3E2723", border_color="#A1887F")
                        signal_entry.pack(pady=(5, 15), padx=15, anchor="w")
                        self.signal_refs[subfolder_key] = signal_entry
                                    
                        def show_suggestions(event, key=subfolder_key, entry=signal_entry, container=frame_container):
                            typed = entry.get()

                            # Create suggestion frame only once
                            if not hasattr(entry, "suggestion_frame") or not entry.suggestion_frame.winfo_exists():
                                entry.suggestion_frame = ctk.CTkFrame(container,width=500, border_width=1, border_color="#AAA",fg_color="transparent")
                                entry.suggestion_frame.pack(padx=15, pady=(0, 15))

                            # Clear previous suggestions
                            for widget in entry.suggestion_frame.winfo_children():
                                widget.destroy()

                            # Hide frame if input is empty
                            if not typed.strip():
                                entry.suggestion_frame.pack_forget()
                                return

                            # Get trie for this subfolder key
                            trie = self.tries.get(key)
                            if not trie:
                                entry.suggestion_frame.pack_forget()
                                return

                            matches = trie.search(typed)

                            if matches:
                                # Show the frame again (in case it was hidden)
                                entry.suggestion_frame.pack(padx=15, pady=(0, 15), fill="x")

                                for name in matches:
                                    def make_cmd(n=name):
                                        def cmd():
                                            entry.delete(0, "end")
                                            entry.insert(0, n)
                                            entry.suggestion_frame.pack_forget()
                                        return cmd

                                    btn = ctk.CTkButton(entry.suggestion_frame, text=name, width=450, fg_color="#E0E0E0",
                                                        hover_color="#C8C8C8", text_color="#000", font=("Arial", 12),
                                                        command=make_cmd())
                                    btn.pack(anchor="w", pady=2, padx=3)
                            else:
                                label = ctk.CTkLabel(entry.suggestion_frame, text="No matches found", text_color="gray")
                                label.pack(anchor="w", pady=5, padx=5)
                        
                        def make_on_first_click(key, entry, container):
                            def on_first_click(event):
                                entry.unbind("<Button-1>")
                                show_suggestions(event, key, entry, container)
                                entry.bind("<KeyRelease>", lambda e: show_suggestions(e, key, entry, container))
                                entry.bind("<FocusOut>", lambda e: getattr(entry, "suggestion_frame", None) and entry.suggestion_frame.pack_forget())
                            return on_first_click

                        signal_entry.bind("<Button-1>", make_on_first_click(subfolder_key, signal_entry, frame_container))
                        
                        def make_on_enter(entry_widget, frames_list, frame_container, mdf, mf4, asc, subfolder_key):
                            def on_enter(event, entry=entry_widget):

                                for widget in frame_container.winfo_children():
                                    if isinstance(widget, (ctk.CTkLabel, ctk.CTkButton)):
                                        text = getattr(widget, "cget", lambda x: "")("text") if hasattr(widget, "cget") else ""
                                        if any(kw in text for kw in ["✅", "❌", "View Plot", "Offset"]):
                                            widget.destroy()
                                
                                frame_container.update()            
                                
                                # Hide suggestion box
                                if hasattr(entry, "suggestion_frame") and entry.suggestion_frame.winfo_exists():
                                    entry.suggestion_frame.pack_forget()

                                # Remove typing events and focus (to remove blinking cursor)
                                entry.unbind("<KeyRelease>")
                                entry.unbind("<FocusOut>")
                                frame_container.focus_set()
                                
                                reference_signal = entry.get().strip()
                                if reference_signal == "":
                                    error_label = ctk.CTkLabel(
                                        frame_container,
                                        text="❌ Please enter a reference signal name.",
                                        font=("Arial", 13),
                                        text_color="red"
                                    )
                                    error_label.pack(padx=15, pady=(5, 0), anchor="w")
                                    return

                                if self.selected_type_of_recording == "CANalyser Plot/Recording":
                                    if mf4 and asc and mdf:
                                        time_shift = calculate_time_difference_between_mf4_and_asc(mf4, asc)
                                        if time_shift is not None:
                                            synchronized_signals_mdf = process_all_signals_mdf(mdf, time_shift)
                                            if synchronized_signals_mdf:
                                                offset, fig1, fig2 = calculate_offset_for_synchronization(mdf, mf4, can_matrix_path,synchronized_signals_mdf, reference_signal)

                                                if offset is not None:
                                                    synchronized_signals_mf4 = process_all_signals_mf4(mf4,offset)
                                                    for frame_key in frames_list:
                                                        self.frame_sync_results[frame_key] = {
                                                            "subfolder": subfolder_key,
                                                            "mf4_file": mf4,
                                                            "mdf_file": mdf,
                                                            "asc_file": asc,
                                                            "offset": offset,
                                                            "time_shift" : time_shift,
                                                            "plot_fig1": fig1,
                                                            "plot_fig2": fig2,
                                                            "synchronized_signals_mdf": synchronized_signals_mdf,
                                                            "synchronized_signals_mf4": synchronized_signals_mf4,}
                                                    processed_subfolders[subfolder_key] = {
                                                        "mf4_file": mf4,
                                                        "mdf_file": mdf,
                                                        "asc_file": asc,
                                                        "offset": offset,
                                                        "time_shift" : time_shift,
                                                        "plot_fig1": fig1,
                                                        "plot_fig2": fig2,
                                                        "synchronized_signals_mdf": synchronized_signals_mdf,
                                                        "synchronized_signals_mf4": synchronized_signals_mf4,}

                                                    sync_label = ctk.CTkLabel(frame_container,
                                                        text=f"✅ Synced Successfully. Offset: {offset:.2f} sec",
                                                        font=("Arial", 13, "bold"), text_color="green")
                                                    sync_label.pack(padx=15, pady=(5, 0), anchor="w")

                                                    view_plot_button = ctk.CTkButton(frame_container, text="View Plot", width=150,
                                                                                     fg_color="#007acc", hover_color="#005f99", text_color="white",
                                                                                     command=lambda f=frames_list[0]: self.show_plot(f))
                                                    view_plot_button.pack(padx=15, pady=(5, 5), anchor="w")
                                                    
                                                    signal_entry.bind("<Button-1>", make_on_first_click(subfolder_key, signal_entry, frame_container))
                                                else:
                                                    error_label = ctk.CTkLabel(frame_container,
                                                        text="❌ Offset could not be calculated. Please try with another reference signal",
                                                        font=("Arial", 13), text_color="red")
                                                    error_label.pack(padx=15, pady=(5, 0), anchor="w")
                                                    signal_entry.bind("<Button-1>", make_on_first_click(subfolder_key, signal_entry, frame_container))
                            return on_enter

                        # ✅ Only one binding now:
                        signal_entry.bind("<Return>", make_on_enter(signal_entry, frames[:], frame_container, mdf, mf4, asc, subfolder_key))
            
            except Exception as e:
                print("Error loading CAN Matrix or processing frames:", e)
        else:
            error_label = ctk.CTkLabel(scrollable_frame, text="CAN Matrix file not found or invalid path.", 
                                    font=("Arial", 14), text_color="red")
            error_label.pack(pady=20)
        
        page.pack(fill="both", expand=True)
        
        ctk.CTkButton(page, text="Back", width=120, fg_color="#99BC85", hover_color="#7a9669",text_color="#000", command=lambda: self.show_page("analysis_scope")
                      ).place(x=40, rely=1.0, y=-50, anchor="sw")

        # Place Next button at bottom-right
        ctk.CTkButton(page, text="Next", width=120, fg_color="#99BC85", hover_color="#7a9669",text_color="#000", command=lambda: self.show_page("processing")
        ).place(relx=1.0, rely=1.0, x=-40, y=-50, anchor="se")
        
        self.add_footer(page)
    
    def show_plot(self, frame_key):
        # Retrieve figures from stored results
        frame_result = self.frame_sync_results.get(frame_key)
        if not frame_result:
            print(f"No plot data found for frame: {frame_key}")
            return

        fig1 = frame_result.get("plot_fig1")
        fig2 = frame_result.get("plot_fig2")

        # Hide the sync page
        if "sync" in self.pages:
            self.pages["sync"].pack_forget()

        # Destroy previous plot page if it exists (optional cleanup)
        if "plot" in self.pages:
            self.pages["plot"].destroy()

        # Create new plot page
        self.pages["plot"] = ctk.CTkFrame(self, fg_color="white")

        title = ctk.CTkLabel(
            self.pages["plot"],
            text=f"Synchronized Plots for Frame: {frame_key}",
            font=("Arial", 16, "bold"),
            text_color="#333"
        )
        title.pack(pady=(15, 5))

        # Plot 1
        canvas1 = FigureCanvasTkAgg(fig1, master=self.pages["plot"])
        canvas1.draw()
        canvas1.get_tk_widget().pack(pady=10)

        # Plot 2
        canvas2 = FigureCanvasTkAgg(fig2, master=self.pages["plot"])
        canvas2.draw()
        canvas2.get_tk_widget().pack(pady=10)

        # Back Button
        self.back_button = ctk.CTkButton(
            self.pages["plot"],
            text="Back",
            width=120,
            fg_color="#99BC85",
            hover_color="#7a9669",
            text_color="#000",
            command=self.show_sync_page
        )
        self.back_button.pack(pady=(10, 20))

        # Show the plot page
        self.pages["plot"].pack(fill="both", expand=True)

    def show_sync_page(self):
        # Hide plot page if visible
        if "plot" in self.pages:
            self.pages["plot"].pack_forget()

        # Show sync page again
        if "sync" in self.pages:
            self.pages["sync"].pack(fill="both", expand=True)
    
    def extract_signal_names_for_subfolder(self, subfolder_path):
    
        signal_names = set()

        mdf_file_path = None
        for f in os.listdir(subfolder_path):
            if f.lower().endswith(".mdf"):
                mdf_file_path = os.path.join(subfolder_path, f)
                break

        # Get the single MF4 file path or None
        mf4_file_path = None
        for f in os.listdir(subfolder_path):
            if f.lower().endswith(".mf4"):
                mf4_file_path = os.path.join(subfolder_path, f)
                break
        
        if mdf_file_path:
            try:
                mdf = mdfreader.Mdf(mdf_file_path)
                signal_names.update(mdf.keys())
            except Exception as e:
                print(f"Error reading MDF file {mdf_file_path}: {e}")
        if mf4_file_path:
            try:
                mf4 = mdfreader.Mdf(mf4_file_path)
                signal_names.update(mf4.keys())
            except Exception as e:
                print(f"Error reading MDF file {mf4_file_path}: {e}")

        return sorted(signal_names)
    
    
    # --------------------------------------LOADING ANIMATION PART---------------------------------------------------------------
    #----------------------------------------------------------------------------------------------------------

    def show_loading_animation(self, parent_window=None):
        parent_window = parent_window if parent_window else self

        # Destroy existing animation if already shown
        if hasattr(self, "loading_frame") and self.loading_frame is not None:
            self.loading_frame.destroy()

        # Create new frame for loading animation
        self.loading_frame = tk.Frame(parent_window)
        self.loading_frame.place(relx=0.5, rely=0.5, anchor="center")  # Center the frame

        self.loading_text_label = tk.Label(self.loading_frame, text="", font=("Arial", 24))
        self.loading_text_label.pack()

        # Set base text depending on function status
        if self.function_status == "Measurement Analysis":
            self.loading_base_text = "LOADING"
        elif self.function_status == "Report Generation":
            self.loading_base_text = "GENERATING REPORT"
        else:
            self.loading_base_text = "PROCESSING"

        # Initialize animation state
        self.loading_current_index = 0
        self.loading_dots = 0

        self.animate_loading()

    
    def animate_loading(self):
        try:
            if hasattr(self, 'loading_text_label') and self.loading_text_label.winfo_exists():
                if self.loading_current_index < len(self.loading_base_text):
                    current_text = self.loading_base_text[:self.loading_current_index + 1]
                    self.loading_text_label.config(text=current_text)
                    self.loading_current_index += 1
                else:
                    dots = '.' * (self.loading_dots % 4)
                    self.loading_text_label.config(text=self.loading_base_text + dots)
                    self.loading_dots += 1

                # Schedule next call and save the id
                self._after_id = self.loading_frame.after(300, self.animate_loading)
        except AttributeError:
            print("")

    def hide_loading_animation(self):
        # Cancel any scheduled animation loop
        if hasattr(self, "_after_id") and self._after_id:
            try:
                self.after_cancel(self._after_id)
            except Exception:
                pass
            self._after_id = None

        # Safely destroy loading frame if it exists and is not already destroyed
        if hasattr(self, "loading_frame") and self.loading_frame:
            try:
                if self.loading_frame.winfo_exists():  # Check if widget still exists
                    self.loading_frame.destroy()
            except Exception as e:
                print(f"⚠️ Error while destroying loading_frame: {e}")
            finally:
                self.loading_frame = None

        # Clear reference to the loading label
        if hasattr(self, "loading_text_label"):
            self.loading_text_label = None
    
    
    
    # -----------------------------------------------------------------------------------------------------
    #----------------------------------------------------------------------------------------------------------
    
    def init_processing_page(self):
        page = ctk.CTkFrame(self, fg_color="#E8DFCA")
        self.pages["processing"] = page

        # Add title based on function_status
        if self.function_status == "Measurement Analysis":
            ctk.CTkLabel(page, text="⚙️ Processing Summary", font=ctk.CTkFont("Arial", 20, "bold"),
                        text_color="#000").pack(pady=(30, 10))
        elif self.function_status == "Report Generation":
            ctk.CTkLabel(page, text="⚙️ Report Generation", font=ctk.CTkFont("Arial", 20, "bold"),
                        text_color="#000").pack(pady=(30, 10))

        if self.function_status == "Report Generation":
            report_template_path = self.saved_files.get("report_path", "")
            if report_template_path:
                self.report_workbook = load_workbook(report_template_path)
                self.report_path_original = report_template_path  # To use later while saving
                print("📄 Report template loaded.")
            else:
                self.report_workbook = None
                print("⚠️ No report template path provided.")
        else:
            self.report_workbook = None  # Clear in other modes

        # Initialize frames to process
        self.frames_to_process = list(self.selected_frames)  # Ensure it's a list
        self.current_frame_index = 0

        # Start processing the frames sequentially
        self.process_next_frame_sequentially()
        
        # Back Button
        button_frame = ctk.CTkFrame(page, fg_color="#E8DFCA")
        button_frame.pack(fill="x", side="bottom", pady=10)
        ctk.CTkButton(button_frame, text="Back to Input", width=120, fg_color="#99BC85", hover_color="#7a9669",
                  text_color="#000", command=lambda: self.show_page("analysis_scope")).pack(side="left", padx=40, anchor="sw")

        # Add footer
        self.add_footer(page)
    
    def process_next_frame_sequentially(self):
        if self.current_frame_index >= len(self.frames_to_process):
            print("✅ All frames processed.")
            self.hide_loading_animation()

            processing_page = self.pages.get("processing")

            if self.function_status == "Report Generation" and self.report_workbook:
                if hasattr(self, "report_path_original") and self.report_path_original:
                    original_dir = os.path.dirname(self.report_path_original)
                    report_save_path = os.path.join(original_dir, "Generated_Report.xlsx")
                    self.report_workbook.save(report_save_path)
                    print(f"✅ Report saved successfully at {report_save_path}")
                    self.report_workbook = None  # Clear after saving

            if processing_page:
                # Common success label
                success_label = ctk.CTkLabel(
                    processing_page,
                    text="✅ All frames processed successfully!",
                    font=("Arial", 14, "bold"),
                    fg_color="#B2F2BB",
                    text_color="black",
                    width=600
                )
                success_label.pack(pady=(10, 10))

                # Additional message based on mode
                if self.function_status == "Report Generation":
                    success_label = ctk.CTkLabel(
                        processing_page,
                        text=f"✅ Report Generated Successfully!\nReport saved at:\n{report_save_path}",
                        font=("Arial", 14, "bold"),
                        fg_color="#B2F2BB",
                        text_color="black",
                        width=600
                    )
                    success_label.pack(pady=(10, 10))
                elif self.function_status == "Measurement Analysis":
                    success_label = ctk.CTkLabel(
                        processing_page,
                        text="🔍 For detailed analysis of each frame, click on Signal Analysis!",
                        font=("Arial", 14, "bold"),
                        fg_color="#B2F2BB",
                        text_color="black",
                        width=600
                    )
                    success_label.pack(pady=(10, 10))
            return

        frame = str(self.frames_to_process[self.current_frame_index])
        print(f"🔄 Processing frame: {frame}")

        def background_processing():
            try:
                self.show_loading_animation()
                self.process_single_frame(frame)

                processing_page = self.pages.get("processing")

                if processing_page and self.function_status == "Report Generation":
                    frame_success_label = ctk.CTkLabel(processing_page,text=f"✅ Frame '{frame}' processed successfully!",font=("Arial", 12),fg_color="#D3F9D8",  # Light green
                                                       text_color="black",width=600)
                    frame_success_label.pack(pady=(2,2))

                self.current_frame_index += 1
                self.after(50, self.process_next_frame_sequentially)

            except Exception as e:
                self.hide_loading_animation()
                tk.messagebox.showerror("Processing Error", f"❌ Failed to process frame {frame}:\n{str(e)}")

        threading.Thread(target=background_processing, daemon=True).start()


    # -----------------------------MEASUREMENT ANALYSIS------------------------------------------------------------------------
    #----------------------------------------------------------------------------------------------------------
    
    def process_single_frame(self, frame):
        
        can_matrix_path = self.saved_files.get("can_matrix", "")
        dtc_matrix_path = self.saved_files.get("dtc_matrix","")
        gateway_sheet_path = self.saved_files.get("gateway_sheet", "")
        
        frame_data = self.frame_sync_results.get(frame, {})

        mf4_file_path = frame_data.get("mf4_file", "")
        mdf_file_path = frame_data.get("mdf_file", "")
        self.offset = frame_data.get("offset", None)
        self.time_shift = frame_data.get("time_shift", None)
        self.synchronized_signals_mf4 = frame_data.get("synchronized_signals_mf4", {})
        self.synchronized_signals_mdf = frame_data.get("synchronized_signals_mdf", {})
        
        # Check if required session values exist
        if self.tx_df.empty or self.rx_df.empty or not mf4_file_path or not mdf_file_path :
            tk.messagebox.showwarning("Missing Files", "Uploaded files are not available. Please go back and upload them.")
            self.show_page("sync")
            return

        if self.tx_df is None or self.tx_df.empty or self.rx_df is None or self.rx_df.empty:
            self.show_page("sync")
            return

        if self.offset is None:
            self.show_page("sync")
            return

        can_matrix_df = pd.concat([self.tx_df, self.rx_df], ignore_index=True)
            
        dbc_signals, asw_signals, frame_id, periodicity, gateway_signals,node,transmitter,receiver = fetch_signals_by_frame_name(can_matrix_df, frame)
        self.dbc_signals=dbc_signals
        self.asw_signals=asw_signals
        self.frame_id=frame_id
        self.periodicity = periodicity
        self.gateway_signals = gateway_signals
        self.node = node
        self.transmitter = transmitter
        self.receiver = receiver
        
        if not dbc_signals and not asw_signals:
            tk.messagebox.showwarning("Frame Not Found", f"The frame '{frame}' is not available in the CAN Matrix.")
            return  # Skips processing this frame without breaking the loop

        synchronized_signals_mf4 = process_all_signals_of_frame(dbc_signals, asw_signals, self.offset,mdf_file_path, mf4_file_path,
                can_matrix_path, self.synchronized_signals_mdf,self.synchronized_signals_mf4)
                
        self.synchronized_signals_mf4 = synchronized_signals_mf4

        transmitter_signals, receiver_signals,signal_type_flag = determine_signal_type(can_matrix_path, frame)
        self.transmitter_signals = transmitter_signals
        self.receiver_signals = receiver_signals
        self.signal_type_flag = signal_type_flag

        if frame_id:
            if dtc_matrix_path:
                matching_signals, mdf = find_error_signals_for_frame(frame, mf4_file_path, self.frame_id)
                self.matching_signals = matching_signals
                self.mdf = mdf

                if matching_signals:
                    activated_signals = search_signals_in_dsm(dtc_matrix_path, self.matching_signals)
                    self.activated_signals = activated_signals

                    if activated_signals:
                        enabled_error_signals = search_enabled_signals_in_excel(
                        dtc_matrix_path, self.activated_signals)
                        self.enabled_error_signals = enabled_error_signals or None
                    else:
                        self.enabled_error_signals = None
                else:
                    self.enabled_error_signals = None
            else:
                self.enabled_error_signals = None

        gateway_types = type_of_gateway(self.gateway_signals, can_matrix_path)
        self.gateway_types = gateway_types

        multiplexor_signals, selector_signals = identify_multiplexor_signals(
                can_matrix_path, self.receiver_signals)
        self.multiplexor_signals = multiplexor_signals
        self.selector_signals = selector_signals

        basic_receiving_signal = extract_basic_receiving_signal(self.receiver_signals, self.multiplexor_signals, self.selector_signals)
        self.basic_receiving_signal = basic_receiving_signal

        print(f"✅ Processed frame: {frame}")

        # Create and collect status DataFrame for each frame
        status_df,signal_info_dict = status_table(
                self.frame_id, self.dbc_signals, self.asw_signals, self.synchronized_signals_mdf,can_matrix_path,
                self.gateway_signals, self.multiplexor_signals, self.selector_signals,
                self.basic_receiving_signal, self.transmitter_signals, self.gateway_types,gateway_sheet_path,mdf_file_path,
                self.synchronized_signals_mf4)
        #status_df["Frame ID"] = self.frame_id
       
        self.status_df = status_df
        self.signal_info_dict = signal_info_dict
        self.frames_data[frame] = {
            "dbc_signals": self.dbc_signals,"asw_signals": self.asw_signals,"frame_id": self.frame_id,
            "periodicity": self.periodicity,"gateway_signals": self.gateway_signals,"signal_type_flag": self.signal_type_flag,
            "synchronized_signals_mf4": self.synchronized_signals_mf4,"synchronized_signals_mdf" : self.synchronized_signals_mdf,
            "transmitter_signals": self.transmitter_signals,"receiver_signals": self.receiver_signals,"enabled_error_signals": self.enabled_error_signals,
            "gateway_types": self.gateway_types,"multiplexor_signals": self.multiplexor_signals,"selector_signals": self.selector_signals,
            "basic_receiving_signal": self.basic_receiving_signal,"status_signal_info_dict": self.signal_info_dict,"time_shift" : self.time_shift,
            "node" : self.node,"transmitter" : self.transmitter,"receiver": self.receiver,"status_dataframe" : self.status_df}
        print("processed suceessfully")
        if self.function_status == "Measurement Analysis":
            self.hide_loading_animation()
            print(status_df)
            frame_widget = self.render_status_table(status_df, frame)
            self.update_idletasks()
            self.update()
            self.frame_widgets.append(frame_widget)
    
        elif self.function_status == "Report Generation":
            print("report")
            self.export_table_to_report(frame)  # Export the report (e.g., to a PowerPoint file)
            self.update_idletasks()  # Update the UI to ensure the report is reflected
            self.update()
    
    
    def render_status_table(self, dataframe, frame_name):
        """Creates a nested scrollable table for each processed frame dynamically."""
    
        # Check if the main scrollable container exists, create if not
        if not hasattr(self, "main_scrollable_container"):
            self.main_scrollable_container = ctk.CTkScrollableFrame(self.pages["processing"], fg_color="#E8DFCA")
            self.main_scrollable_container.pack(pady=10, padx=10, fill="both", expand=True)
            self.processed_frames = []  # Store processed frame names
        # Prevent duplicate frames from being added
        if frame_name in self.processed_frames:
            return
        self.processed_frames.append(frame_name)
        print("rendering")
        # Create a separate frame inside the main scrollable container
        table_container = ctk.CTkFrame(self.main_scrollable_container,fg_color="#E8DFCA",border_width=2,corner_radius=10)
        table_container.pack(pady=10, fill="x", padx=5)

        # Title for processed frame
        title_label = tk.Label(table_container,text=f"Processed Frame: {frame_name}",font=("Arial", 14, "bold"),bg="#E8DFCA",fg="#5D4037",anchor="w")
        title_label.pack(pady=5, padx=10, anchor="w")

        # Create an inner scrollable frame for the table
        table_frame = ctk.CTkScrollableFrame(table_container, width=800, height=150, fg_color="#f9f9f9")
        table_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        style = ttk.Style()

        # Apply overall theme (Windows or Clam is most reliable)
        style.theme_use("clam")  # or 'default' on macOS

        # Configure Treeview background
        style.configure("Custom.Treeview",
            background="#FBE9E7",        # row background
            foreground="#3E2723",        # row text
            rowheight=60,
            fieldbackground="#FBE9E7",
            font=("Arial", 11))

        # Configure heading
        style.configure("Custom.Treeview.Heading",
            background="#D7CCC8",
            foreground="#5D4037",       # header text color
            font=("Arial", 12, "bold"))

        # Optional: highlight row on hover or selection
        style.map("Custom.Treeview",
            background=[("selected", "#A1887F")],
            foreground=[("selected", "white")])
        
        # Create a Treeview Table inside the scrollable frame
        columns = list(dataframe.columns)
        max_rows = 5
        visible_rows = min(len(dataframe), max_rows)

        tree = ttk.Treeview(table_frame,columns=columns,show="headings",height=visible_rows,style="Custom.Treeview")

        # Define column headings
        for i, col in enumerate(columns):
            tree.heading(col, text=col, anchor="center")
            if i == 0:  # First column
                tree.column(col, anchor="center", width=150)
            elif i == 1:  # Second column
                tree.column(col, anchor="center", width=200)
            elif i == 2:  # Third column
                tree.column(col, anchor="center", width=150)
            else:  # Remaining columns
                tree.column(col, anchor="center", width=100)

        # Insert rows into the table
        for _, row in dataframe.iterrows():
            tree.insert("", "end", values=list(row))

        # Add scrollbars
        scrollbar_y = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        scrollbar_y.pack(side="right", fill="y")
        tree.configure(yscrollcommand=scrollbar_y.set)

        scrollbar_x = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
        scrollbar_x.pack(side="bottom", fill="x")
        tree.configure(xscrollcommand=scrollbar_x.set)
    

        tree.pack(expand=True, fill="both")
        action_frame = ctk.CTkFrame(table_container, fg_color="transparent")
        action_frame.pack(fill="x", pady=(10, 5), padx=5)

        # Status message
        if not dataframe.empty:
            message_text = f"✅ {frame_name} processed successfully "
            message_color = "#2E7D32"  # Green
        else:
            message_text = f"{frame_name} failed to process ❌"
            message_color = "#C62828"  # Red

        status_label = tk.Label(
            action_frame,
            text=message_text,
            font=("Arial", 14, "bold"),
            fg=message_color,
            bg="#E8DFCA",  # match container
            anchor="w"
        )
        status_label.pack(side="left")

        # Signal Analysis button
        next_button = ctk.CTkButton(
            action_frame,
            text="Signal Analysis",font=("Arial", 14, "bold"),
            command=lambda fn=frame_name: threading.Thread(
                target=self.signal_analysis_handler,
                args=(fn,)
            ).start()
        )
        next_button.pack(side="right")
        self.update_idletasks() 
        return table_container
    
    def signal_analysis_handler(self, frame_name):
        frame_id = "0x" + frame_name.split('_')[-1].rstrip('h')
        self.open_frame_window(frame_name, frame_id)
    
    def display_figures_in_scrollable_window(self, figures, additional_figures, parent_frame):
        if not isinstance(figures, list):
            figures = [figures]
        if not isinstance(additional_figures, list):
            additional_figures = [additional_figures]

        # Create scrollable canvas
        canvas = tk.Canvas(parent_frame)
        scrollbar = ttk.Scrollbar(parent_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Add all figures
        all_figures = figures + additional_figures
        for fig in all_figures:
            fig_canvas = FigureCanvasTkAgg(fig, master=scrollable_frame)
            fig_widget = fig_canvas.get_tk_widget()
            fig_widget.pack(pady=10, fill="both", expand=True)

            fig.set_size_inches(6, 4)

    def open_frame_window(self, frame_name, frame_id):
        new_window = tk.Toplevel(self)
        new_window.title(f"Signal Analysis - {frame_name}")
        new_window.geometry("800x600")

        frame_datas = self.frame_sync_results.get(frame_name, {})

        mf4_file_path = frame_datas.get("mf4_file", "")
        mdf_file_path = frame_datas.get("mdf_file", "")
        self.offset = frame_datas.get("offset", None)
        self.time_shift = frame_datas.get("time_shift", None)
        self.synchronized_signals_mf4 = frame_datas.get("synchronized_signals_mf4", {})
        self.synchronized_signals_mdf = frame_datas.get("synchronized_signals_mdf", {})
        
        can_matrix_path = self.saved_files.get("can_matrix", "")
        dtc_matrix_path = self.saved_files.get("dtc_matrix", "")
        gateway_sheet = self.saved_files.get("gateway_sheet", "")
        fid_mapping_path = self.saved_files.get("fid_sheet", "")

        title = tk.Label(new_window, text=f"Frame: {frame_name}", font=("Arial", 14, "bold"))
        title.pack(pady=10)
    
        result_frame = ttk.Frame(new_window)
        result_frame.pack(fill="both", expand=True, padx=10, pady=10)
    
        print(f"Starting processing for frame: {frame_name}")

        def run_quality_check():
            try:
            # 📦 Fetch data from frames_data dictionary
                frame_data = self.frames_data.get(frame_name)

                if not frame_data:
                    raise ValueError(f"No data found for frame: {frame_name}")

                print("Starting backend processing with frame data...")
                self.show_loading_animation(new_window)
                #🔥 Call your backend function with the frame-specific data
                df,color_dict,figure_dict = quality_status_of_signal(frame_id,frame_data["dbc_signals"],frame_data["asw_signals"],frame_data["synchronized_signals_mf4"],
                frame_data["enabled_error_signals"],frame_data["gateway_signals"],frame_data["status_signal_info_dict"],frame_data["gateway_types"],
                frame_data["synchronized_signals_mdf"],frame_data["basic_receiving_signal"],frame_data["multiplexor_signals"],frame_data["selector_signals"],
                frame_data["transmitter_signals"],frame_name,dtc_matrix_path,mf4_file_path,can_matrix_path,gateway_sheet,fid_mapping_path,mdf_file_path,self.offset,
                self.time_shift,self.selected_templates_by_type)

                print("Processing complete.")
                self.qulaity_analysis_dataframe = df
                self.frame_name = frame_name
                self.quality_analysis_page_window = result_frame
                self.frames_data[frame_name].update({"qulaity_analysis_dataframe": df,"frame_name": frame_name,"quality_analysis_page_window": result_frame})
                self.hide_loading_animation()
                self.after(0, self.render_status_table_in_window, df, frame_name, result_frame, color_dict,figure_dict)

            except Exception as e:
                self.hide_loading_animation()
                error_message = f"Error: {str(e)}"

        # Run backend function in a new thread so UI doesn't freeze
        thread = threading.Thread(target=run_quality_check, daemon=True)
        thread.start()
        print("Thread started.")

        # Back button to close the window
        tk.Button(new_window, text="Back", command=new_window.destroy).pack(pady=10)
    
    def render_status_table_in_window(self, dataframe, frame_name, parent_frame, color_dict, figure_dict):
        print(f"Rendering table for frame: {frame_name}")  # Debug print

        self.figure_dict = figure_dict

        # Title label for the frame
        title_label = tk.Label(parent_frame, text=f"Signal Analysis: {frame_name}",
                            font=("Arial", 14, "bold"))
        title_label.pack(pady=10)

        # Create a container frame for the table
        status_table_frame = tk.Frame(parent_frame)
        status_table_frame.pack(fill="both", expand=True)

        table_frame = tk.Frame(status_table_frame)
        table_frame.pack(fill="both", expand=True)

        data = dataframe.values.tolist()
        columns = dataframe.columns.tolist()

        bg_colors = {}

        for row_index in range(len(dataframe)):
            for col_index in [3, 4, 6]:  # Status columns
                value = dataframe.iloc[row_index, col_index]
                if value == "       ╔═══╗\n       ║   ✔ ║\n       ╚═══╝":
                    bg_colors[(row_index, col_index)] = "#c6f6c3"
                elif value == "       ╔═══╗\n       ║   – ║\n       ╚═══╝":
                    bg_colors[(row_index, col_index)] = "#d3d3d3"
                elif value == "       ╔═══╗\n       ║   ✘ ║\n       ╚═══╝":
                    bg_colors[(row_index, col_index)] = "#f6c3c3"

            # Determine signal name
            network_signal = dataframe.iloc[row_index, 1]
            if "Upstream Signal:" in network_signal:
                signal = network_signal.split("Upstream Signal:")[1].split("\n")[0].strip()
            else:
                signal = network_signal.strip()

            if color_dict:
                basic_color = color_dict.get(signal, {}).get("basic_replacement_status")
                gateway_color = color_dict.get(signal, {}).get("signal_gateway_sub_status")

                if basic_color:
                    bg_colors[(row_index, 5)] = basic_color
                    dataframe.iloc[row_index, 5] = ""
                if gateway_color:
                    bg_colors[(row_index, 7)] = gateway_color
                    dataframe.iloc[row_index, 7] = ""

        # Create a fresh Sheet instance per frame
        sheet = Sheet(table_frame,
                    data=data,
                    headers=columns,
                    height=500,
                    width=1000,
                    header_font=("Arial", 11, "bold"),
                    font=("Arial", 12, "normal"),
                    show_x_scrollbar=True,
                    show_y_scrollbar=True,
                    header_width=200,
                    header_height=70)

        for row in range(len(data)):
            sheet.row_height(row, height=70)

        for col in range(len(columns)):
            if col == 0 or col == 2:
                sheet.column_width(col, width=220)
            elif col == 1:
                sheet.column_width(col, width=480)
            else:
                sheet.column_width(col, width=170)

        sheet.pack(fill="both", expand=True)
        sheet.enable_bindings(('single'))

        # Highlight relevant cells
        for (r, c), color in bg_colors.items():
            sheet.highlight_cells(row=r, column=c, bg=color, fg="black", redraw=False)

        sheet.redraw()
        print(type(sheet))
        # Bind clicks if needed
        self.bind_table_clicks(sheet,frame_name)

        # Optionally store for reference (if you want to access it later)
        if not hasattr(self, "frame_table_refs"):
            self.frame_table_refs = {}
        self.frame_table_refs[frame_name] = {
            "status_table_frame": status_table_frame,
            "sheet": sheet
        }


    def bind_table_clicks(self, sheet, frame_name):
        """Bind table click events to the provided Sheet instance."""
        sheet.bind("<ButtonRelease-1>", lambda event: self.cell_clicked(sheet, frame_name, event))
        sheet.bind("<ButtonRelease-2>", lambda event: self.cell_clicked(sheet, frame_name, event))
        sheet.bind("<ButtonRelease-3>", lambda event: self.cell_clicked(sheet, frame_name, event))

    def cell_clicked(self,sheet,frame_name, event=None):
        print("entered")
        currently_selected = sheet.get_currently_selected()
        if currently_selected:
            row, col = currently_selected[0], currently_selected[1]
            status = sheet.get_cell_data(row, col)

            if col in [5,7] and status in "    ┏━━━━━━┓\n    ┃  View Plots    ┃\n    ┗━━━━━━┛":
                self.on_click(sheet,frame_name, event=None)

            if col in [3, 4, 6] and status in [
                "       ╔═══╗\n       ║   ✔ ║\n       ╚═══╝",
                "       ╔═══╗\n       ║   ✘ ║\n       ╚═══╝",
                "       ╔═══╗\n       ║   – ║\n       ╚═══╝"
            ]:
                print("clicked")
                cell_content = sheet.get_cell_data(row, 1)
                dbc_signal = ""
                upstream_signal = ""
                downstream_signal = ""

                if "Upstream Signal:" in cell_content:
                    upstream_signal = cell_content.split("Upstream Signal:")[1].split("\n")[0].strip()
                    if "Downstream Signal:" in cell_content:
                        downstream_signal = cell_content.split("Downstream Signal:")[1].split("\n")[0].strip()
                    dbc_signal = f"{upstream_signal} -> {downstream_signal}"
                else:
                    dbc_signal = cell_content

                if dbc_signal in self.selector_signals:
                    return

                if col == 4:
                    asw_signal = sheet.get_cell_data(row, 2)
                elif col == 6:
                    asw_signal = downstream_signal
                elif col == 3:
                    signal_info = self.status_signal_info_dict.get(upstream_signal)
                    asw_signal = signal_info.get("downstream_signal") or signal_info.get("downstream_signal_frame") if signal_info else None

                self.on_icon_click(dbc_signal, asw_signal, status, row, col,
                                 frame_name, self.figure_dict, sheet)

    def on_icon_click(self, dbc_signal, asw_signal, status, row_index, col, frame_name, figure_dict, sheet):
        print(f"{dbc_signal}      {asw_signal}     {status}")
        frame_data = self.frames_data.get(frame_name)
        if not frame_data:
            raise ValueError(f"No data found for frame: {frame_name}")

        synchronized_signals_mf4 = frame_data["synchronized_signals_mf4"]
        synchronized_signals_mdf = frame_data["synchronized_signals_mdf"]

        can_matrix_path = self.saved_files.get("can_matrix", "")
        mdf_file_path = self.saved_files.get("mdf_file", "")

        boxed_text = "    ┏━━━━━━┓\n    ┃  View Plots    ┃\n    ┗━━━━━━┛"

        if status in ["       ╔═══╗\n       ║   – ║\n       ╚═══╝"]:
            if "->" in dbc_signal:
                signal1, signal2 = map(str.strip, dbc_signal.split("->"))
            else:
                signal1 = dbc_signal
            popup = tk.Toplevel()
            popup.title("Communication Successful")
            popup.geometry("800x800")
            popup.attributes('-topmost', True)

            if col == 4:
                tk.Label(popup, text=f"The current value and replacement value are same", fg="green", font=("Arial", 12, "bold")).pack(pady=5)
                figures = plot_communication_sucessful(signal1, asw_signal, synchronized_signals_mf4, "basic", self.offset, mdf_file_path, can_matrix_path)

            elif col == 6:
                tk.Label(popup, text=f"The current value and gateway substitution value are same", fg="green", font=("Arial", 12, "bold")).pack(pady=5)
                figures = plot_communication_sucessful(signal1, asw_signal, synchronized_signals_mdf, "signal_gateway", self.offset, mdf_file_path, can_matrix_path)

            elif col == 3:
                tk.Label(popup, text="Communication Successful for all other timestamps.\nHighlight region: Since there is no valid upstream samples there is no transmission in the downstream side", fg="green", font=("Arial", 12, "bold")).pack(pady=5)
                figures = plot_communication_sucessful(signal1, asw_signal, synchronized_signals_mdf, "frame_gateway", self.offset, mdf_file_path, can_matrix_path)

            if figures:
                self.display_figures_in_scrollable_window(figures, [], popup)

            tk.Button(popup, text="Back", command=popup.destroy, font=("Arial", 12)).pack(pady=10)

        elif status in ["       ╔═══╗\n       ║   ✔ ║\n       ╚═══╝"]:
            signal1 = dbc_signal.split("->")[0].strip() if "->" in dbc_signal else dbc_signal
            popup = tk.Toplevel()
            popup.title("Communication Successful")
            popup.geometry("800x800")
            popup.attributes('-topmost', True)

            tk.Label(popup, text=f"Communication Successful {signal1} vs {asw_signal}", fg="green", font=("Arial", 12, "bold")).pack(pady=5)

            if col == 3:
                figures = plot_communication_sucessful(signal1, asw_signal, synchronized_signals_mdf, "frame_gateway", self.offset, mdf_file_path, can_matrix_path)
            elif col == 6:
                figures = plot_communication_sucessful(signal1, asw_signal, synchronized_signals_mdf, "signal_gateway", self.offset, mdf_file_path, can_matrix_path)
            elif col == 4:
                figures = plot_communication_sucessful(signal1, asw_signal, synchronized_signals_mf4, "basic", self.offset, mdf_file_path, can_matrix_path)

            if figures:
                self.display_figures_in_scrollable_window(figures, [], popup)

            tk.Button(popup, text="Back", command=popup.destroy, font=("Arial", 12)).pack(pady=10)

        elif status in ["       ╔═══╗\n       ║   ✘ ║\n       ╚═══╝"]:
            signal1 = dbc_signal.split("->")[0].strip() if "->" in dbc_signal else dbc_signal

            if col == 4:
                sheet.set_cell_data(row_index, 5, boxed_text, redraw=True)
                sheet.update()
                self._register_clickable_cell(row_index, 5, signal1, figure_dict,sheet,frame_name)

            elif col == 6:
                sheet.set_cell_data(row_index, 7, boxed_text, redraw=True)
                sheet.update()
                self._register_clickable_cell(row_index, 7, signal1, figure_dict,sheet,frame_name)
    

    def _register_clickable_cell(self, row, col, signal, figure_dict,sheet,frame_name):
        print("Registering clickable cell at", row, col)
        self.clickable_cells[(row, col)] = signal
        self.figure_dict = figure_dict  # Store for access in on_click
        sheet.bind("<ButtonRelease-1>", lambda event: self.on_click(sheet,frame_name, event))  # Use bound method
    
    def on_click(self,sheet,frame_name, event=None):
        selected = sheet.get_currently_selected()
        if not selected:
            return
        clicked_row, clicked_col = selected[0], selected[1]
        key = (clicked_row, clicked_col)

        if key in self.clickable_cells:
            signal = self.clickable_cells[key]
            print(f"Clicked on cell {key} for signal '{signal}'")
            print(self.figure_dict)

            if clicked_col == 5 and signal in self.figure_dict:
                fig = self.figure_dict[signal].get("basic_communication_figure", [])
                figs = fig[0] if isinstance(fig, list) and len(fig) == 1 and isinstance(fig[0], list) else fig

                def show_popup():
                    popup = tk.Toplevel()
                    popup.title("Communication Successful")
                    popup.geometry("800x800")
                    popup.attributes('-topmost', True)

                    self.current_win = popup
                    self.window_open = True

                    

                    label = tk.Label(popup,
                                    text="Communication Successful for all other timestamps.\nHighlight region: Since there is no valid upstream samples there is no transmission in the downstream side",
                                    fg="green", font=("Arial", 12, "bold"))
                    label.pack(pady=5)

                    self.display_figures_in_scrollable_window(figs, [], popup)

                    # ✅ Use proper back handler
                    back_button = tk.Button(popup, text="Back", command=self.on_back_button_pressed(sheet,frame_name), font=("Arial", 12))
                    back_button.pack(pady=10)

                if len(figs) == 1:
                    sheet.after(0, show_popup)
                else:
                    self.display_figures_in_new_window(figs,sheet,frame_name)

            elif clicked_col == 7 and signal in self.figure_dict:
                fig = self.figure_dict[signal].get("signal_gateway_figure", [])
                figs = fig[0] if isinstance(fig, list) and len(fig) == 1 and isinstance(fig[0], list) else fig

                def show_popup():
                    popup = tk.Toplevel()
                    popup.title("Communication Successful")
                    popup.geometry("800x800")
                    popup.attributes('-topmost', True)

                    self.current_win = popup
                    self.window_open = True

                    label = tk.Label(popup,
                                    text="Communication Successful for all other timestamps.\nHighlight region: Since there is no valid upstream samples there is no transmission in the downstream side",
                                    fg="green", font=("Arial", 12, "bold"))
                    label.pack(pady=5)

                    self.display_figures_in_scrollable_window(figs, [], popup)

                    # ✅ Use proper back handler
                    back_button = tk.Button(popup, text="Back", command=self.on_back_button_pressed(sheet,frame_name), font=("Arial", 12))
                    back_button.pack(pady=10)

                if len(figs) == 1:
                    sheet.after(0, show_popup)
                else:
                    self.display_figures_in_new_window(figs,sheet,frame_name)
    
    def display_figures_in_new_window(self, figures,sheet,frame_name):
        if not figures:
            print("No figures to display.")
            return

        if self.window_open and self.current_win:
            self.current_win.destroy()
            self.window_open = False

        new_win = tk.Toplevel()
        new_win.geometry("900x600")
        new_win.title("Plots for Analysis")
        self.window_open = True
        self.current_win = new_win

        # Handle user closing window with the X button
        def on_close():
            new_win.after(0, self.on_back_button_pressed(sheet,frame_name))

        new_win.protocol("WM_DELETE_WINDOW", on_close)

        canvas = tk.Canvas(new_win)
        scrollbar = tk.Scrollbar(new_win, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        def render_figures(fig_list):
            if not fig_list:
                return

            # First figure (full width)
            first_fig = fig_list[0]
            first_canvas = FigureCanvasTkAgg(first_fig, master=scrollable_frame)
            first_canvas.draw()
            first_canvas.get_tk_widget().pack(fill="both", expand=True, pady=10)

            # Remaining figures in grid
            remaining_figs = fig_list[1:]
            if remaining_figs:
                grid_frame = tk.Frame(scrollable_frame)
                grid_frame.pack(pady=10)

                for idx, fig in enumerate(remaining_figs):
                    fig_canvas = FigureCanvasTkAgg(fig, master=grid_frame)
                    fig_canvas.draw()
                    widget = fig_canvas.get_tk_widget()
                    row = idx // 2
                    col = idx % 2
                    widget.grid(row=row, column=col, padx=10, pady=10, sticky="nsew")
                    grid_frame.grid_rowconfigure(row, weight=1)
                    grid_frame.grid_columnconfigure(col, weight=1)

        render_figures(figures)

        back_button = tk.Button(new_win,text="Back",command=lambda: new_win.after(0, self.on_back_button_pressed(sheet,frame_name)),font=("Arial", 12, "bold"))
        back_button.pack(pady=10)
    
    def on_back_button_pressed(self, sheet=None, frame_name=None):
        # Close popup window if open
        if self.window_open and self.current_win:
            self.current_win.destroy()
            self.window_open = False

        # Show the status_table_frame for the given frame_name if available
        if frame_name and hasattr(self, "frame_table_refs") and frame_name in self.frame_table_refs:
            ref = self.frame_table_refs[frame_name]
            status_table_frame = ref.get("status_table_frame")
            if status_table_frame and status_table_frame.winfo_exists():
                status_table_frame.pack(fill="both", expand=True)
                status_table_frame.lift()
                status_table_frame.focus_set()

        # Focus the passed sheet and bind clicks
        if sheet:
            sheet.focus_set()
            self.bind_table_clicks(sheet, frame_name)
        
    
    
    # -----------------------------REPORT GENERATION------------------------------------------------------------------------
    #----------------------------------------------------------------------------------------------------------------------
    
    def extend_rows_with_layout(self, sheet, layout_start_row, layout_row_count, total_signals):
        """
        Extend rows for the specific table block only.
        Assumes table block starts at layout_start_row and currently has layout_row_count rows.
        """
        # Number of rows currently allocated for this table block
        new_start_of_second_table,end_of_second_table = 18,19
        current_block_rows = layout_row_count
        
        
        # If total signals fit in the current block, no extension needed
        if total_signals <= current_block_rows:
            return new_start_of_second_table,end_of_second_table 

        # Calculate how many rows to add
        rows_to_add = total_signals - current_block_rows
        
        # Insert empty rows just below the current table block to make room
        insert_at_row = layout_start_row + current_block_rows
        sheet.insert_rows(insert_at_row, amount=rows_to_add)

        # Copy layout rows repeatedly into the newly inserted rows
        for i in range(rows_to_add):
            new_row_num = insert_at_row + i
            layout_row_to_copy = layout_start_row + (i % layout_row_count)

            # Copy cells and styles from layout row
            for cell in sheet[layout_row_to_copy]:
                new_cell = sheet.cell(row=new_row_num, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.alignment = cell.alignment.copy()
                    new_cell.number_format = cell.number_format
                    new_cell.protection = cell.protection.copy()
                
                if not isinstance(new_cell, type(sheet.merged_cells)):
                    new_cell.value = cell.value
            # Copy merged cells for that specific layout row only (IMPORTANT: use layout_row_to_copy, new_row_num)
            merged_ranges = list(sheet.merged_cells.ranges)
            for merged_range in merged_ranges:
                min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
                # Only copy merged ranges that are in the layout row being copied
                if min_row == layout_row_to_copy and max_row == layout_row_to_copy:
                    start_col = get_column_letter(min_col)
                    end_col = get_column_letter(max_col)
                    new_range = f"{start_col}{new_row_num}:{end_col}{new_row_num}"
                    try:
                        sheet.merge_cells(new_range)
                    except Exception as e:
                        print(f"Error merging cells {new_range}: {e}")
            
        next_row = layout_start_row + current_block_rows + rows_to_add
        self.clear_row_style(sheet, next_row)  # Call using self
        print(f"🧹 Cleared styles on row {next_row} to avoid half-border effect.")
        
        last_row_of_block = insert_at_row + rows_to_add - 1
        desired_gap = 2

        print(f"📌 First table ends at row {last_row_of_block}")

        # Dynamically detect the second table start
        next_table_start_row = last_row_of_block + 1
        while True:
            if any(cell.value for cell in sheet[next_table_start_row]):
                break
            next_table_start_row += 1

        print(f"📌 Detected second table starting at row {next_table_start_row}")

        # Adjust the second table position
        template_header_row = 18
        template_body_row = 19
        template_rows_to_copy = [template_header_row, template_body_row]

        second_table_row_count = 2
        new_start_of_second_table = last_row_of_block + desired_gap + 1
        print(f"📌 Second table should start at row {new_start_of_second_table}")

        # 1. Save cell styles and values from the current location (actual second table)
        saved_cells = {}
        for row in range(next_table_start_row, next_table_start_row + second_table_row_count):
            saved_cells[row] = []
            for cell in sheet[row]:
                saved_cells[row].append({
                    'column': cell.column,
                    'value': cell.value,
                    'font': cell.font.copy(),
                    'fill': cell.fill.copy(),
                    'border': cell.border.copy(),
                    'alignment': cell.alignment.copy(),
                    'number_format': cell.number_format,
                    'protection': cell.protection.copy(),
                })

        # 2. Get merged ranges from template rows only
        merged_ranges_to_shift = []
        for merged_range in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            if min_row in template_rows_to_copy or max_row in template_rows_to_copy:
                merged_ranges_to_shift.append(str(merged_range))

        # 3. Delete and insert rows at new position
        sheet.delete_rows(next_table_start_row, amount=second_table_row_count)
        sheet.insert_rows(new_start_of_second_table, amount=second_table_row_count)

        # 4. Paste values and styles to new location
        for old_row, cells in saved_cells.items():
            new_row = new_start_of_second_table + (old_row - next_table_start_row)
            for cell_info in cells:
                new_cell = sheet.cell(row=new_row, column=cell_info['column'])
                new_cell.value = cell_info['value']
                new_cell.font = cell_info['font']
                new_cell.fill = cell_info['fill']
                new_cell.border = cell_info['border']
                new_cell.alignment = cell_info['alignment']
                new_cell.number_format = cell_info['number_format']
                new_cell.protection = cell_info['protection']

        # 5. Reapply merged cells using template logic (not from original second table)
        for merged_range in merged_ranges_to_shift:
            min_col, min_row, max_col, max_row = range_boundaries(merged_range)
            row_offset = new_start_of_second_table - template_header_row
            new_min_row = min_row + row_offset
            new_max_row = max_row + row_offset
            new_range = f"{get_column_letter(min_col)}{new_min_row}:{get_column_letter(max_col)}{new_max_row}"
            try:
                sheet.merge_cells(new_range)
            except Exception as e:
                print(f"❌ Merge failed for {new_range}: {e}")

        print(f"✅ Second table is now correctly positioned at row {new_start_of_second_table} with template styling.")

        layout_start_row = new_start_of_second_table + 1  # e.g., 23
        layout_row_count = 1  # Only 1 row in layout for second table
        insert_at_row = layout_start_row + 1  # Insert below the template row
        rows_to_add = total_signals - 1  # One row already exists, add the rest
        
        if rows_to_add > 0:
            print(f"📌 Extending second table with {rows_to_add} rows using styled copy")

            for i in range(rows_to_add):
                new_row_num = insert_at_row + i
                layout_row_to_copy = layout_start_row + (i % layout_row_count)

                for cell in sheet[layout_row_to_copy]:
                    new_cell = sheet.cell(row=new_row_num, column=cell.column, value=cell.value)
                    if cell.has_style:
                        new_cell.font = cell.font.copy()
                        new_cell.fill = cell.fill.copy()
                        new_cell.border = cell.border.copy()
                        new_cell.alignment = cell.alignment.copy()
                        new_cell.number_format = cell.number_format
                        new_cell.protection = cell.protection.copy()

                # Handle merged cells in the template row
                merged_ranges = list(sheet.merged_cells.ranges)
                for merged_range in merged_ranges:
                    min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
                    if min_row == layout_row_to_copy and max_row == layout_row_to_copy:
                        start_col = get_column_letter(min_col)
                        end_col = get_column_letter(max_col)
                        new_range = f"{start_col}{new_row_num}:{end_col}{new_row_num}"
                        try:
                            sheet.merge_cells(new_range)
                        except Exception as e:
                            print(f"⚠️ Error merging cells {new_range}: {e}")
            
            end_of_second_table = insert_at_row + rows_to_add - 1
            print(f"✅ Second table extended down to row {insert_at_row + rows_to_add - 1}")
        
        return new_start_of_second_table,end_of_second_table
    
    
    def clear_row_style(self, sheet, row: int):
        for cell in sheet[row]:
            cell.font = Font()              # default font
            cell.fill = PatternFill(fill_type=None)  # no fill
            cell.border = Border()          # no border
            cell.alignment = Alignment()    # default alignment
            cell.number_format = 'General'  # reset number format
            cell.protection = Protection()  # default.
    

    def export_table_to_report(self, frame_name):
        print("Generating Report in Thread:", threading.current_thread().name)
        workbook = self.report_workbook  # Main sheet (first sheet)
        main_sheet = workbook.worksheets[0]

        # Debug: Print existing sheet values
        print("\n🔍 Existing Values in Sheet:")
        for row in main_sheet.iter_rows(values_only=True):
            print(row)

        # Check for the first empty row in 'FRAME NAME' column (D)
        row = 2
        frame_row_mapping = {}  
        while True:
            cell = main_sheet[f'D{row}']
            if not cell.value or str(cell.value).strip() == "":
                print(f"✅ Empty row found at Row {row} - Inserting Frame Name: {frame_name}")
                cell.value = frame_name
                main_sheet[f'A{row}'].value = row - 1  # Serial number (1, 2, 3...)
                
                frame_data = self.frames_data.get(frame_name)
                frame_id = frame_data["frame_id"]
                node = frame_data["node"]
                periodicity = frame_data["periodicity"]
                transmitter = frame_data["transmitter"]
                receiver = frame_data["receiver"]
                signal_type = frame_data["signal_type_flag"]
                # Set the Frame ID as a hyperlink
                main_sheet[f'C{row}'].value = node
                main_sheet[f'E{row}'].value = signal_type
                frame_id_cell = main_sheet[f'B{row}']
                frame_id_cell.value = frame_id
                frame_id_cell.hyperlink = f"#{quote_sheetname(frame_id)}!A1"
                frame_id_cell.style = "Hyperlink"
                frame_row_mapping[frame_name] = row  # Save the row number here
                break
            row += 1

        # Create or get the sheet for the Frame ID
        if frame_id in workbook.sheetnames:
            frame_sheet = workbook[frame_id]
        else:
            # Copy the layout from Sheet1 to the new sheet
            template_sheet = workbook["Detail_Sheet_Layout"]
            frame_sheet = workbook.create_sheet(title=frame_id)
            self.copy_sheet_layout(template_sheet, frame_sheet)
        
        frame_sheet['B1'] = frame_name  # Frame Name
        frame_sheet['B2'] = frame_id 
        frame_sheet['B3'] = periodicity
        frame_sheet['B4'] = transmitter
        frame_sheet['B5'] = receiver  
        
        dbc_signals = frame_data["dbc_signals"]
        asw_signals = frame_data["asw_signals"]
        signal_count = len(set(dbc_signals))
        
        frame_datas = self.frame_sync_results.get(frame_name, {})

        mf4_file_path = frame_datas.get("mf4_file", "")
        mdf_file_path = frame_datas.get("mdf_file", "")
        self.offset = frame_datas.get("offset", None)
        self.time_shift = frame_datas.get("time_shift", None)
        self.sub_folder = frame_datas.get("subfolder",None)
        
        can_matrix_path = self.saved_files.get("can_matrix", "")
        dtc_matrix_path = self.saved_files.get("dtc_matrix", "")
        gateway_sheet = self.saved_files.get("gateway_sheet", "")
        fid_mapping_path = self.saved_files.get("fid_sheet", "")
        
        df,color_dict,figure_dict = quality_status_of_signal(frame_id,frame_data["dbc_signals"],frame_data["asw_signals"],frame_data["synchronized_signals_mf4"],
                frame_data["enabled_error_signals"],frame_data["gateway_signals"],frame_data["status_signal_info_dict"],frame_data["gateway_types"],
                frame_data["synchronized_signals_mdf"],frame_data["basic_receiving_signal"],frame_data["multiplexor_signals"],frame_data["selector_signals"],
                frame_data["transmitter_signals"],frame_name,dtc_matrix_path,mf4_file_path,can_matrix_path,gateway_sheet,fid_mapping_path,mdf_file_path,self.offset,
                self.time_shift,self.selected_templates_by_type)
        
        # Start filling SL NO and signal columns from row 10
        print("quality status table created")
        start_row = 10
        layout_row_count = 1  # Use this row for layout reference
        
        second_table_start_row,second_table_end_row = self.extend_rows_with_layout(frame_sheet, start_row, layout_row_count, signal_count)
        print("extended")
        used_dbc_signals = set()  # Track seen dbc_signals
        row = start_row

        for dbc_signal, asw_signal in zip(dbc_signals, asw_signals):
            cleaned_dbc = dbc_signal.strip().upper()
            if cleaned_dbc in used_dbc_signals:
                continue  # Skip duplicates

            used_dbc_signals.add(cleaned_dbc)

            sl_no = row - start_row + 1
            frame_sheet[f'A{row}'] = sl_no
            frame_sheet[f'C{row}'] = dbc_signal
            frame_sheet[f'E{row}'] = asw_signal

            df.columns = df.columns.str.strip()
            
            # Step 1: Extract effective signal from the 'Network Signal' column
            df["Effective Signal"] = df["Network Signal"].apply(
                lambda x: next((line.split(":", 1)[1].strip().upper()
                                for line in str(x).splitlines()
                                if "Upstream Signal:" in line),
                            str(x).strip().upper()))

            # Step 2: Find matching rows based on Effective Signal
            matching_rows = df[df["Effective Signal"] == cleaned_dbc]

            # Step 3: Extract Node values
            node_series = matching_rows["Node"]

            if not node_series.empty:
                node_lines = node_series.dropna().astype(str).unique()
                combined_node = "\n".join(node_lines)

                cell = frame_sheet[f'B{row}']
                cell.value = combined_node
                cell.alignment = Alignment(wrap_text=True)
            else:
                frame_sheet[f'B{row}'] = ""

            row += 1  # Move to next row only if a new (unique) dbc_signal was processed
        
        total_rows_written = row - start_row
        original_total_rows = signal_count

        extra_rows = original_total_rows - total_rows_written
        if extra_rows > 0:
            frame_sheet.delete_rows(row, extra_rows)
        print("✅ First table is filled with unique DBC signals successfully.")
        
        frame_sheet.row_dimensions[second_table_start_row].height = 120

        all_ok = True
        red_hex = "ffc6c3"
        second_table_data_start = second_table_start_row + 1
        
        for j in range(signal_count):
            try:
                row_num = second_table_data_start + j
                sl_no = j + 1
                frame_sheet[f'A{row_num}'] = sl_no
                
                # Column B: Node (wrap text)
                node_cell = frame_sheet[f'B{row_num}']
                node_cell.value = df.loc[j, "Node"]
                node_cell.alignment = Alignment(wrap_text=True)
                
                # Column C (merged with D): Network Signal (wrap text)
                network_cell = frame_sheet[f'C{row_num}']
                network_cell.value = df.loc[j, "Network Signal"]
                network_cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                 
                # Column E (merged with F): ASW Signal (wrap text)
                asw_cell = frame_sheet[f'E{row_num}']
                # Safely get "ASW Signal" value, fallback to empty string if missing or NaN
                asw_signal_value = df.loc[j, "ASW Signal"] if "ASW Signal" in df.columns else ""
                if pd.isna(asw_signal_value):
                    asw_signal_value = ""

                asw_cell.value = asw_signal_value
                asw_cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                bold_font = Font(bold=True)

                def get_icon_from_box(value):
                    tick_box = "       ╔═══╗\n       ║   ✔ ║\n       ╚═══╝"
                    cross_box = "       ╔═══╗\n       ║   ✘ ║\n       ╚═══╝"
                    dash_box = "       ╔═══╗\n       ║   – ║\n       ╚═══╝"

                    # Make sure value is string and strip whitespace
                    if not isinstance(value, str):
                        value = str(value) if value is not None else ""

                    value = value.strip()

                    if value == tick_box.strip():
                        return '✔', bold_font
                    elif value == cross_box.strip():
                        return '✘', bold_font
                    elif value == dash_box.strip():
                        return '-', bold_font
                    elif value == "--":
                        return "--", None
                    else:
                        return '', None

                # BASIC COMMUNICATION STATUS (I)
                basic_comm_status = df.loc[j, "Basic Communication\n Status"] if "Basic Communication\n Status" in df.columns else "--"
                if pd.isna(basic_comm_status):
                    basic_comm_status = "--"
                icon, font = get_icon_from_box(str(basic_comm_status))
                cell = frame_sheet[f'I{row_num}']
                cell.value = icon if icon else "--"
                if font: 
                    cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='center')
               
                # FRAME GATEWAY STATUS (L)
                frame_gateway_status = df.loc[j, "Frame Gateway\nStatus"] if "Frame Gateway\nStatus" in df.columns else "--"
                if pd.isna(frame_gateway_status):
                    frame_gateway_status = "--"
                icon, font = get_icon_from_box(str(frame_gateway_status))
                cell = frame_sheet[f'L{row_num}']
                cell.value = icon if icon else "--"
                if font:
                    cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # SIGNAL GATEWAY STATUS (M)
                signal_gateway_status = df.loc[j, "Signal Gateway\nStatus"] if "Signal Gateway\nStatus" in df.columns else "--"
                if pd.isna(signal_gateway_status):
                    signal_gateway_status = "--"
                icon, font = get_icon_from_box(str(signal_gateway_status))
                cell = frame_sheet[f'M{row_num}']
                cell.value = icon if icon else "--"
                if font:
                    cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                def get_color_info(signal_name):
                    if not isinstance(signal_name, str):
                        return None

                    # Direct match
                    color_info = color_dict.get(signal_name.strip())
                    if color_info:
                        return color_info

                    # If it's a compound signal, extract upstream
                    if "Upstream Signal:" in signal_name and "Downstream Signal:" in signal_name:
                        try:
                            lines = signal_name.split("\n")
                            upstream_line = [line for line in lines if "Upstream Signal:" in line]
                            if upstream_line:
                                upstream_signal = upstream_line[0].split("Upstream Signal:")[1].strip()
                                return color_dict.get(upstream_signal)
                        except Exception:
                            return None
                    return None

                network_signal = network_cell.value
                color_info = get_color_info(network_signal)

                # === Column J: basic_replacement_status ===
                j_cell = frame_sheet[f'J{row_num}']
                j_color_code = None

                if color_info and 'basic_replacement_status' in color_info:
                    j_color_code = color_info['basic_replacement_status']
                    if j_color_code:
                        fill = PatternFill(start_color=j_color_code.replace("#", ""), end_color=j_color_code.replace("#", ""), fill_type="solid")
                        j_cell.fill = fill

                is_cell_filled = (j_cell.fill is not None and j_cell.fill.fill_type is not None and
                                j_cell.fill.start_color is not None and j_cell.fill.start_color.rgb not in ("00000000", "FFFFFFFF", None))

                if (j_cell.value is None or str(j_cell.value).strip() == "") and not is_cell_filled:
                    j_cell.value = "--"
                
                
                # === Column N: signal_gateway_sub_status ===
                n_cell = frame_sheet[f'N{row_num}']
                n_color_code = None

                if color_info and 'signal_gateway_sub_status' in color_info:
                    n_color_code = color_info['signal_gateway_sub_status']
                    if n_color_code:
                        fill = PatternFill(start_color=n_color_code.replace("#", ""), end_color=n_color_code.replace("#", ""), fill_type="solid")
                        n_cell.fill = fill

                is_cell_filled = (n_cell.fill is not None and n_cell.fill.fill_type is not None and
                                n_cell.fill.start_color is not None and n_cell.fill.start_color.rgb not in ("00000000", "FFFFFFFF", None))

                if (n_cell.value is None or str(n_cell.value).strip() == "") and not is_cell_filled:
                    n_cell.value = "--"
               
                
                print("filled signal successfully in the table")
            except Exception as e:
                print(f"Error processing row {j}: {e}")
                import traceback
                traceback.print_exc()
                break  # optional, to stop after the error for debugging
            
        
        print("Second table is filled successfully")
        green_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")  # ForestGreen
        red_fill = PatternFill(start_color="B22222", end_color="B22222", fill_type="solid")    # FireBrick
        
        def insert_figure(worksheet, figure, cell_location):
            try:
                if figure is None:
                    print("No figure provided, skipping insertion.")
                    return False  # or just `return` if you don't need a value

                img_buffer = BytesIO()
                figure.savefig(img_buffer, format='png', bbox_inches='tight')
                plt.close(figure)
                img_buffer.seek(0)
                img = OpenpyxlImage(img_buffer)
                worksheet.add_image(img, cell_location)
                return True

            except Exception as e:
                print(f"Error adding figure: {e}")
                raise

        def insert_heading(worksheet, text, row, font_size=10):
            cell = worksheet[f"A{row}"]
            cell.value = text
            cell.font = Font(bold=True, size=font_size, color="000000")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            worksheet.merge_cells(f"A{row}:C{row}")

            # Set fill and font color conditionally
            if text.startswith("Network Signal:"):
                fill_color = "4B0082"  # Dark Purple
                font_color = "FFFFFF"  # White
            elif "Basic Communication" in text or "Signal Gateway" in text:
                fill_color = "FFFF00"  # Yellow
                font_color = "000000"  # Black
            else:
                fill_color = None
                font_color = "000000"  # Default black text

            # Apply formatting to merged cells
            if fill_color:
                for col in range(ord('A'), ord('C') + 1):
                    col_letter = chr(col)
                    cell_ref = f"{col_letter}{row}"
                    worksheet[cell_ref].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    worksheet[cell_ref].font = Font(bold=True, size=font_size, color=font_color)

            return row + 1

        def insert_single_figure(worksheet, figure, row, height=35):
            cell_location = f"A{row}"
            inserted = insert_figure(worksheet, figure, cell_location)
            return row + height if inserted else row

        def insert_remaining_figures_grid(worksheet, figures, start_row, figures_per_row=3, col_spacing=6, row_height=30):
            col_index = 0
            current_row = start_row
            for fig in figures:
                if fig:
                    col_letter = chr(ord('A') + col_index * col_spacing)
                    cell_location = f"{col_letter}{current_row}"
                    inserted = insert_figure(worksheet, fig, cell_location)
                    if inserted:
                        col_index += 1
                        if col_index == figures_per_row:
                            col_index = 0
                            current_row += row_height
            if col_index >= 0:
                current_row += row_height
            return current_row

        # === Start inserting figures after second table ===
        strat_current_row = second_table_end_row + 4

        for network_signal in list(set(dbc_signals)):
            if network_signal not in figure_dict:
                continue

            figures = figure_dict[network_signal]
            if not figures:
                continue

            # --- Ensure all figure categories are lists ---
            if 'basic_communication_figure' in figures and not isinstance(figures['basic_communication_figure'], list):
                figures['basic_communication_figure'] = [figures['basic_communication_figure']]
            if 'signal_gateway_figure' in figures and not isinstance(figures['signal_gateway_figure'], list):
                figures['signal_gateway_figure'] = [figures['signal_gateway_figure']]

            # --- Network Signal Heading ---
            current_row = insert_heading(frame_sheet, f"Network Signal: {network_signal}", strat_current_row, font_size=12)
            print(f"after heading the current row is :{current_row}")
            current_row += 1  # Leave 3-row space
            
            def flatten_figures(fig_list):
                flat = []
                for f in fig_list:
                    if isinstance(f, list):
                        flat.extend(flatten_figures(f))
                    else:
                        flat.append(f)
                return flat


            # --- Basic Communication Section ---
            if 'basic_communication_figure' in figures and figures['basic_communication_figure']:
                basic_figs = flatten_figures(figures['basic_communication_figure'])
                print(basic_figs)
                current_row = insert_heading(frame_sheet, "Basic Communication Figures", current_row, font_size=10)
                current_row += 2  # Leave 2-row space

                # Insert first figure
                current_row = insert_single_figure(frame_sheet, basic_figs[0], current_row)
                print(f"after inserting the first figure the current row is :{current_row}")

                # Insert remaining figures in 3x3 layout
                if len(basic_figs) > 1:
                    print(f"remaining figure- starts inserting at the current row is :{current_row}")
                    current_row = insert_remaining_figures_grid(frame_sheet, basic_figs[1:], current_row)
                    print(f"after inserting the remaining figure the current row is :{current_row}")

            # --- Signal Gateway Section ---
            if 'signal_gateway_figure' in figures and figures['signal_gateway_figure']:
                gateway_figs = flatten_figures(figures['signal_gateway_figure'])
                print(gateway_figs)
                current_row = insert_heading(frame_sheet, "Signal Gateway Figures", current_row, font_size=10)
                print(f"after inserting the signal heading the current row is :{current_row}")
                current_row += 2  # Leave 2-row space

                # Insert first figure
                current_row = insert_single_figure(frame_sheet, gateway_figs[0], current_row)
                print(f"after inserting the first figure the current row is :{current_row}")

                # Insert remaining figures in 3x3 layout
                if len(gateway_figs) > 1:
                    current_row = insert_remaining_figures_grid(frame_sheet, gateway_figs[1:], current_row)
                    print(f"after inserting the remaining figure the current row is :{current_row}")

            # --- Leave space before next signal ---
            current_row += 1
            print(f"second signal at row  :{current_row}")
            strat_current_row = current_row

        print("Figures inserted properly")

        # Update status in main sheet
        row_in_main_sheet = frame_row_mapping.get(frame_name)
        if row_in_main_sheet:
            main_sheet[f'F{row_in_main_sheet}'].fill = green_fill if all_ok else red_fill
            main_sheet[f'G{row_in_main_sheet}'].value = self.sub_folder
        
        
    
    def copy_sheet_layout(self, source_sheet, target_sheet):
        """
        Copy the layout (formatting, merged cells, column widths) from the source sheet to the target sheet.
        """
        # Copy row heights
        for row in source_sheet.row_dimensions:
            try:
                target_sheet.row_dimensions[row].height = source_sheet.row_dimensions[row].height
            except Exception as e:
                print(f"Error copying row height for row {row}: {e}")

        # Copy column widths
        for col in source_sheet.column_dimensions:
            try:
                target_sheet.column_dimensions[col].width = source_sheet.column_dimensions[col].width
            except Exception as e:
                print(f"Error copying column width for column {col}: {e}")

        # Copy cell values and formatting
        for row in source_sheet.iter_rows():
            for cell in row:
                try:
                    new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)

                    if cell.has_style:
                        new_cell.font = cell.font.copy()
                        new_cell.fill = cell.fill.copy()
                        new_cell.border = cell.border.copy()
                        new_cell.alignment = cell.alignment.copy()
                        new_cell.number_format = cell.number_format
                        new_cell.protection = cell.protection.copy()
                except Exception as e:
                    print(f"Error copying cell at {cell.coordinate}: {e}")

        # Copy merged cells
        for merged_range in source_sheet.merged_cells.ranges:
            try:
                target_sheet.merge_cells(str(merged_range))
            except Exception as e:
                print(f"Error merging cells {merged_range}: {e}")

    
    # ------------------------------------------------------------------------------------------------------------------------
    #------------------------------------------------------------------------------------------------------------------------


    def validate_and_go_to_sync(self):
        can_path = self.can_entry.get().strip()
        dtc_path = self.dtc_entry.get().strip()
        gateway_path = self.gateway_entry.get().strip()
        fid_path = self.fid_entry.get().strip()
        folder_path = self.folder_entry.get().strip()

        # Validate only the required file paths (CAN and folder, and valid folder flag)
        if not all([can_path, folder_path, self.folder_valid]):
            tk.messagebox.showwarning("Missing Input", "Please load all the required files properly (CAN Matrix and Folder).")
            return

        try:
            # Ensure temp directory exists
            os.makedirs(self.temp_dir, exist_ok=True)

            # Copy required CAN file
            can_dest = shutil.copy(can_path, os.path.join(self.temp_dir, os.path.basename(can_path)))

            # Copy optional files only if provided
            dtc_dest = shutil.copy(dtc_path, os.path.join(self.temp_dir, os.path.basename(dtc_path))) if dtc_path else None
            gateway_dest = shutil.copy(gateway_path, os.path.join(self.temp_dir, os.path.basename(gateway_path))) if gateway_path else None
            fid_dest = shutil.copy(fid_path, os.path.join(self.temp_dir, os.path.basename(fid_path))) if fid_path else None

            # Save paths for further use (excluding mdf, mf4, asc files)
            self.saved_files = {
                "folder_path": folder_path,  # Save only the folder path now
                "can_matrix": can_dest,
                "dtc_matrix": dtc_dest,
                "gateway_sheet": gateway_dest,
                "fid_sheet": fid_dest
            }

            if self.function_status.lower() == "report generation":
                report_path = self.report_entry.get().strip()
                if report_path:
                    self.saved_files["report_path"] = report_path
            self.sync_status = "success"
            self.show_page("type_of_project")

        except Exception as e:
            print("Error saving files:", e)
            self.sync_status = "error"
            
    def show_page(self, name):
        print(name)
        if name not in self.pages:
            return  # If the page doesn't exist, do nothing
        current_page = None
        loading_label = None
       
        for page in self.pages.values():
            if page.winfo_ismapped():  # Find the currently displayed page
                current_page = page
                break

        if current_page:
            button_widgets = [widget for widget in current_page.winfo_children() if isinstance(widget, ctk.CTkButton)]

            if button_widgets:
                # Find the button with text "Next"
                next_button = None
                for btn in button_widgets:
                    if btn.cget("text") == "Next":
                        next_button = btn
                        break
                print(f"next button :{next_button}")
                if next_button:
                    # Place "Loading..." label right below the "Next" button
                    loading_label = tk.Label(current_page,text="Loading.....",font=("Arial", 14),fg="black",bg=current_page["bg"])
                    loading_label.place(x=next_button.winfo_x(),y=next_button.winfo_y() + next_button.winfo_height() + 5)
                    self.update_idletasks()
        
        # Recreate the page from scratch
        if name == "sync":
            self.init_sync_page()
        elif name == "analysis_scope":
            self.init_analysis_scope_page()
        elif name == "processing":
            self.init_processing_page()
        elif name == "type_of_project":
            print("entering")
            self.init_template_page()
      
        
        # Show the new page
        for page in self.pages.values():
            page.pack_forget()  # Hide all pages

        if loading_label:
            loading_label.destroy()
       
    
        self.pages[name].pack(fill="both", expand=True)


        
    def select_can(self):
        """Handle CAN Matrix file selection."""
        file = filedialog.askopenfilename()
        if file:
            self.can_entry.delete(0, "end")
            self.can_entry.insert(0, file)
            print(f"Selected CAN Matrix Path: {file}")  # Debugging

    def select_gateway(self):
        """Handle CAN Matrix file selection."""
        file = filedialog.askopenfilename()
        if file:
            self.gateway_entry.delete(0, "end")
            self.gateway_entry.insert(0, file)
            print(f"Selected Gateway Sheet Path: {file}")  # Debugging

    def select_fid(self):
        """Handle CAN Matrix file selection."""
        file = filedialog.askopenfilename()
        if file:
            self.fid_entry.delete(0, "end")
            self.fid_entry.insert(0, file)
            print(f"Selected FID Mapping Path: {file}")  # Debugging

    def select_report_tempelate(self):
        """Handle CAN Matrix file selection."""
        file = filedialog.askopenfilename()
        if file:
            self.report_entry.delete(0, "end")
            self.report_entry.insert(0, file)
            self.report_path_original = file
            print(f"Selected Report Tempelate Path: {file}")  # Debugging

    def select_dtc(self):
        """Handle DTC Matrix file selection."""
        file = filedialog.askopenfilename()
        if file:
            self.dtc_entry.delete(0, "end")
            self.dtc_entry.insert(0, file)
            print(f"Selected DTC Matrix Path: {file}")  # Debugging

    def select_folder(self):
        """Handle Measurement Folder selection."""
        path = filedialog.askdirectory()
        if path:
            self.folder_entry.delete(0, "end")
            self.folder_entry.insert(0, path)
            print(f"Selected Measurement Folder Path: {path}")  # Debugging

            folder = Path(path)
            # Check if there is at least one subfolder
            subfolders = [f for f in folder.iterdir() if f.is_dir()]
            self.folder_valid = len(subfolders) > 0

            # Handle error display
            if self.folder_error_label:
                self.folder_error_label.destroy()
            if not self.folder_valid:
                self.folder_error_label = ctk.CTkLabel(
                    self.pages["measurement"],
                    text="Selected folder should not be empty.",
                    text_color="red"
                )
                self.folder_error_label.pack(pady=(5, 0))
    
    def add_footer(self, frame):
        footer = ctk.CTkFrame(frame, fg_color="#E8DFCA", height=100)
        footer.pack(side="bottom", fill="x")

        try:
            if os.path.exists(self.footer_path):
                footer_img = ctk.CTkImage(Image.open(self.footer_path), size=(900, 20))
                ctk.CTkLabel(footer, image=footer_img, text="").pack(side="top", fill="x", pady=(0, 10))
        except:
            pass

        info = ctk.CTkFrame(footer, fg_color="transparent")
        info.pack(fill="x", padx=20)
        ctk.CTkLabel(info, text="Version: 1.1", font=("Arial", 12), text_color="#000").pack(side="left")
        ctk.CTkLabel(info, text="©CANwiser MS/EPA-PJ5-PS", font=("Arial", 12), text_color="#000").place(relx=0.5, rely=0.5, anchor="center")

        try:
            if os.path.exists(self.logo_path):
                logo = ctk.CTkImage(Image.open(self.logo_path), size=(120, 40))
                ctk.CTkLabel(info, image=logo, text="").pack(side="right")
        except:
            pass


if __name__ == "__main__":
    app = CANwiserApp()
    app.mainloop()